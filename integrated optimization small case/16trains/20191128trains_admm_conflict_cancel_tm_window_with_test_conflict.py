# -*- coding: utf-8 -*-
"""
Created on Wed Aug  7 11:00:10 2019

@author: zhangqin
"""
import xlrd
import pandas as pd
import numpy as np
import openpyxl
import time
import copy
import operator
import random

g_node_list = []
g_train_list = []
g_link_list = []
g_station_list = []
g_blockage_list = []


g_number_of_state=0
g_number_of_nodes = 0
g_number_of_trains = 0
g_number_of_links = 0
g_number_of_stations = 0
g_number_of_blockages = 0

class Station:
    def __init__(self):
        self.station_id = 0
        self.outbound_a_node = 0
        self.outbound_d_node = 0
        self.inbound_a_node = 0
        self.inbound_d_node = 0
        self.outbound_stop_nodes_list=[]
        self.inbound_stop_nodes_list=[]
        self.bothside_stop_nodes_list = []
        self.outbound_nonstop_node = 0
        self.inbound_nonstop_node = 0
        self.outbound_stop_a_links_list = []
        self.outbound_stop_d_links_list = []
        self.outbound_nonstop_a_links_list = []
        self.outbound_nonstop_d_links_list = []
        self.inbound_stop_a_links_list = []
        self.inbound_stop_d_links_list = []
        self.inbound_nonstop_a_links_list = []
        self.inbound_nonstop_d_links_list = []
        self.station_route_run_tm=0
        self.outbound_wait_links_list=[]
        self.inbound_wait_links_list=[]
        self.wait_links_list = []

    
class Node:
    def __init__(self):
        self.node_id = 0
        self.node_type = 0
        self.in_station_number = 0
        self.min_track_dwell_tm = 0
        self.outgoing_node_list = []
        self.outgoing_link_list = []   
        self.ingoing_link_list =[]
        self.IN_after_arcs_of_td_receive_node=[]
        self.IN_before_and_after_arcs_of_td_receive_node=[]
        self.IN_after_arcs_of_td_depart_node=[]
        self.IN_before_and_after_arcs_of_td_depart_node=[]
        self.time_dependent_LR_multiplier=[]
        self.station_route_run_tm=0
        self.tm_already_list=[]
        self.cost_already_list=[]
    
class Link:
    def __init__(self):
        self.link_id = 0
        self.link_type = 0
        self.from_node_id = 0
        self.to_node_id = 0
        self.train_run_tm = 0
        self.link_capacity = 0
        self.link_cost = 0
        self.conflict_link_list = []
        self.time_dependent_LR_multiplier = []
        self.time_dependent_LR_cost = [] 
        self.time_dependent_ADMM_cost = [] 
        self.IN_before_arcs_of_td_link =[]
        self.IN_after_arcs_of_td_link =[]
        self.IN_before_and_after_arcs_of_td_link =[]
        
        
class Blockage:
    def __init__(self):
        self.blockage_id = 0
        self.blockage_node_id = 0
        self.blockage_link_list = []
        self.from_time_interval = 0
        self.to_time_interval = 0
        self.blockage_type= 0
        
        
class Train:
    def __init__(self):
        self.train_id = 0
        self.train_direction = 0 
        self.actual_orig_node_id = 0
        self.actual_dest_node_id = 0
        self.original_tm_beginning= 0
        self.plan_dest_arrival_tm = 0
        self.station_sequence_list = []
        self.stop_pattern_list = []     
        self.plan_station_d_tm_list = []
        self.min_dwell_tm_list = []
        self.max_dwell_tm_list = []
        self.track_preference=[]
        self.available_nodes_list = []
        self.available_links_list = []
        self.available_arcs_list = []
        
        self.Label_cost=0
        self.PrimalLabelCost=0
        self.current_iteration_upper_bound_priority = 0

        self.time_dependent_link_volume = []
        self.time_dependent_implied_waiting_link_flag = []
        self.best_time_dependent_link_volume = []
        
        self.node_sequence = []
        self.time_sequence = []
        self.link_sequence = []
        self.best_node_sequence = []
        self.best_time_sequence = []
        self.best_link_sequence = []
        self.node_sequence_upper_bound = []
        self.time_sequence_upper_bound = []
        self.link_sequence_upper_bound = []
        self.best_node_sequence_upper_bound=[]
        self.best_time_sequence_upper_bound=[]
        self.best_link_sequence_upper_bound=[]
        self.node_sequence_lower_bound = []
        self.time_sequence_lower_bound = []
        self.link_sequence_lower_bound = []
        self.best_node_sequence_lower_bound=[]
        self.best_time_sequence_lower_bound=[]
        self.best_link_sequence_lower_bound=[]
        self.conflict_trains_list = []

        self.train_free_flow_time=0
    
def node_link_map():
    global node_link_map
    node_link_map = pd.DataFrame({'link_id':[],'from_node_id':[],'to_node_id':[]})    
    for l in range(1, g_number_of_links):
        node_link_map = node_link_map.append({'link_id':g_link_list[l].link_id,'from_node_id':g_link_list[l].from_node_id,'to_node_id':g_link_list[l].to_node_id},ignore_index=True)
    return ()    
    
def train_available_nodes_list(train_direction,stop_pattern_list,station_sequence_list):
    train_available_nodes_list = []
    if train_direction == -1:
        for sta_ii in range(0,len(station_sequence_list)):
            station = station_sequence_list[sta_ii]
            train_available_nodes_list.extend(g_station_list[station].outbound_stop_nodes_list)
            train_available_nodes_list.append(g_station_list[station].outbound_a_node)
            train_available_nodes_list.append(g_station_list[station].outbound_d_node)
            if stop_pattern_list[sta_ii]==0:
                train_available_nodes_list.append(g_station_list[station].outbound_nonstop_node)
    else:
        for sta_ii in range(0,len(station_sequence_list)):
            station = station_sequence_list[sta_ii]
            train_available_nodes_list.extend(g_station_list[station].inbound_stop_nodes_list)
            train_available_nodes_list.append(g_station_list[station].inbound_a_node)
            train_available_nodes_list.append(g_station_list[station].inbound_d_node)
            if stop_pattern_list[sta_ii]==0:
                train_available_nodes_list.append(g_station_list[station].inbound_nonstop_node)
    return train_available_nodes_list

def train_available_links_list(train_direction,stop_pattern_list,station_sequence_list,actual_orig_node_id,actual_dest_node_id):
    train_available_links_list = []
    if train_direction == -1:
        for sta_ii in range(0,len(station_sequence_list)):
            station = station_sequence_list[sta_ii]
            # add station links
            train_available_links_list.extend(g_station_list[station].outbound_stop_a_links_list)
            train_available_links_list.extend(g_station_list[station].outbound_stop_d_links_list)
            train_available_links_list.extend(g_station_list[station].outbound_wait_links_list)
            if stop_pattern_list[sta_ii]==0:
                train_available_links_list.append(g_station_list[station].outbound_nonstop_a_links_list)
                train_available_links_list.append(g_station_list[station].outbound_nonstop_d_links_list)
            # add section links
            if sta_ii < len(station_sequence_list)-1:
                station_last=station_sequence_list[sta_ii+1]
                link_pre_n = g_station_list[station].outbound_d_node
                link_after_n = g_station_list[station_last].outbound_a_node
                section_link_id =int (node_link_map.loc[(node_link_map.from_node_id == link_pre_n) & \
                                                        (node_link_map.to_node_id == link_after_n),'link_id'].values[0])
                train_available_links_list.append(section_link_id)
    else:
        for sta_ii in range(0,len(station_sequence_list)):
            station = station_sequence_list[sta_ii]
            # add station links
            train_available_links_list.extend(g_station_list[station].inbound_stop_a_links_list)
            train_available_links_list.extend(g_station_list[station].inbound_stop_d_links_list)
            train_available_links_list.extend(g_station_list[station].inbound_wait_links_list)
            if stop_pattern_list[sta_ii]==0:
                train_available_links_list.append(g_station_list[station].inbound_nonstop_a_links_list)
                train_available_links_list.append(g_station_list[station].inbound_nonstop_d_links_list)
            # add section links
            if sta_ii < len(station_sequence_list)-1:
                station_last=station_sequence_list[sta_ii+1]
                link_pre_n = g_station_list[station].inbound_d_node
                link_after_n = g_station_list[station_last].inbound_a_node
                section_link_id =int (node_link_map.loc[(node_link_map.from_node_id == link_pre_n) & \
                                                        (node_link_map.to_node_id == link_after_n),'link_id'].values[0])
                train_available_links_list.append(section_link_id)
    virtual_orig_wt_link = int (node_link_map.loc[(node_link_map.from_node_id == actual_orig_node_id) & \
                                   (node_link_map.to_node_id == actual_orig_node_id),'link_id'].values[0])
    virtual_dest_wt_link = int (node_link_map.loc[(node_link_map.from_node_id == actual_dest_node_id) & \
                                   (node_link_map.to_node_id == actual_dest_node_id),'link_id'].values[0])
    train_available_links_list.append(virtual_orig_wt_link)
    train_available_links_list.append(virtual_dest_wt_link)           
    return train_available_links_list

def train_available_arcs_list(train_available_links_list,
                              actual_orig_node_id,
                              actual_dest_node_id,
                              original_tm_beginning,
                              plan_station_d_tm_list,
                              plan_dest_arrival_tm):
    train_available_arcs_list =-1*np.ones([g_number_of_links, g_number_of_time_intervals]).astype(int)

    l_orig_waiting=int (node_link_map.loc[(node_link_map.from_node_id == actual_orig_node_id) & \
                                                        (node_link_map.to_node_id == actual_orig_node_id),'link_id'].values[0])
    last_station_depart_tm=plan_station_d_tm_list[-1]
    for t in range(max(1,original_tm_beginning-T_orig_eariler),min(original_tm_beginning+T_orig_later+1,g_number_of_time_intervals)):
        train_available_arcs_list[l_orig_waiting][t]=1
    for l in train_available_links_list:
        if l!=l_orig_waiting and g_link_list[l].to_node_id!=actual_dest_node_id:
            for t in range(max(1,original_tm_beginning-T_orig_eariler),min(last_station_depart_tm+T_dest_later+1,g_number_of_time_intervals)):
                train_available_arcs_list[l][t]=1
                if g_link_list[l].to_node_id==g_link_list[l].from_node_id:
                    for t in range(min(last_station_depart_tm+T_dest_later+1,g_number_of_time_intervals),\
                                   min(last_station_depart_tm+max_t_implicit_occupation+T_dest_later+1,g_number_of_time_intervals)):
                        train_available_arcs_list[l][t]=1
        if g_link_list[l].to_node_id==actual_dest_node_id and g_link_list[l].from_node_id!=g_link_list[l].to_node_id:
            for t in range(max(1,original_tm_beginning-T_orig_eariler),min(last_station_depart_tm+T_dest_later+1,g_number_of_time_intervals)):
                train_available_arcs_list[l][t]=1
        if g_link_list[l].to_node_id==actual_dest_node_id and g_link_list[l].from_node_id==g_link_list[l].to_node_id:
            for t in range(max(1,original_tm_beginning-T_orig_eariler),g_number_of_time_intervals):
                train_available_arcs_list[l][t]=1
    return train_available_arcs_list


def get_sta_ii_of_waiting_node_train_id_ranked(n,k):
    for sta_ii in range(0,len(g_train_list[k].station_sequence_list)):
        station_id=g_train_list[k].station_sequence_list[sta_ii]
        if n in g_station_list[station_id].bothside_stop_nodes_list:
            sta_ii_of_n_in_station_sequence=sta_ii
            break
    return(sta_ii_of_n_in_station_sequence)

                
    
def g_ReadInputData():
    # initialization
    global g_number_of_nodes
    global g_number_of_trains
    global g_number_of_links
    global g_number_of_stations
    global g_number_of_blockages
    global g_all_stations_stop_nodes_list
    global g_all_stations_wait_links_list
    g_all_stations_stop_nodes_list = []
    g_all_stations_wait_links_list = []
    
    #read station data
    book = xlrd.open_workbook("input_station_structure.xlsx")
    sheet = book.sheet_by_index(0)
    station = Station()
    station.station_id = 0
    g_station_list.append(station)
    g_number_of_stations += 1
    for l in range(1,sheet.nrows):
        #try:
            station = Station()
            station.station_id = int(sheet.cell_value(l,0))
            station.outbound_a_node = int(sheet.cell_value(l,1))
            station.outbound_d_node = int(sheet.cell_value(l,2))
            station.inbound_a_node = int(sheet.cell_value(l,3))
            station.inbound_d_node = int(sheet.cell_value(l,4))
            outbound_stop_nodes_list = str(sheet.cell_value(l,5))
            station.outbound_stop_nodes_list = outbound_stop_nodes_list.strip().split(';')
            station.outbound_stop_nodes_list = [float (links) for links in station.outbound_stop_nodes_list]
            station.outbound_stop_nodes_list = [int (links) for links in station.outbound_stop_nodes_list]
            inbound_stop_nodes_list = str(sheet.cell_value(l,6))
            station.inbound_stop_nodes_list = inbound_stop_nodes_list.strip().split(';')
            station.inbound_stop_nodes_list = [float (links) for links in station.inbound_stop_nodes_list]
            station.inbound_stop_nodes_list = [int (links) for links in station.inbound_stop_nodes_list]            
            station.bothside_stop_nodes_list = list(set(station.outbound_stop_nodes_list+station.inbound_stop_nodes_list))            
            g_all_stations_stop_nodes_list.extend(station.bothside_stop_nodes_list)
            station.outbound_nonstop_node = int(sheet.cell_value(l,7))
            station.inbound_nonstop_node = int(sheet.cell_value(l,8))            
            outbound_stop_a_links_list = str(sheet.cell_value(l,9))
            station.outbound_stop_a_links_list = outbound_stop_a_links_list.strip().split(';')
            station.outbound_stop_a_links_list = [float (links) for links in station.outbound_stop_a_links_list]
            station.outbound_stop_a_links_list = [int (links) for links in station.outbound_stop_a_links_list]
            outbound_stop_d_links_list = str(sheet.cell_value(l,10))
            station.outbound_stop_d_links_list = outbound_stop_d_links_list.strip().split(';')
            station.outbound_stop_d_links_list = [float (links) for links in station.outbound_stop_d_links_list]
            station.outbound_stop_d_links_list = [int (links) for links in station.outbound_stop_d_links_list]
            station.outbound_nonstop_a_links_list = int(sheet.cell_value(l,11))
            station.outbound_nonstop_d_links_list = int(sheet.cell_value(l,12))
            inbound_stop_a_links_list = str(sheet.cell_value(l,13))
            station.inbound_stop_a_links_list = inbound_stop_a_links_list.strip().split(';')
            station.inbound_stop_a_links_list = [float (links) for links in station.inbound_stop_a_links_list]
            station.inbound_stop_a_links_list = [int (links) for links in station.inbound_stop_a_links_list]
            inbound_stop_d_links_list = str(sheet.cell_value(l,14))
            station.inbound_stop_d_links_list = inbound_stop_d_links_list.strip().split(';')
            station.inbound_stop_d_links_list = [float (links) for links in station.inbound_stop_d_links_list]
            station.inbound_stop_d_links_list = [int (links) for links in station.inbound_stop_d_links_list]
            station.inbound_nonstop_a_links_list = int(sheet.cell_value(l,15))
            station.inbound_nonstop_d_links_list = int(sheet.cell_value(l,16))            
            station.station_route_run_tm = int(sheet.cell_value(l,17))
            outbound_wait_links_list = str(sheet.cell_value(l,18))
            station.outbound_wait_links_list = outbound_wait_links_list.strip().split(';')
            station.outbound_wait_links_list = [float (links) for links in station.outbound_wait_links_list]
            station.outbound_wait_links_list = [int (links) for links in station.outbound_wait_links_list]
            inbound_wait_links_list = str(sheet.cell_value(l,19))
            station.inbound_wait_links_list = inbound_wait_links_list.strip().split(';')
            station.inbound_wait_links_list = [float (links) for links in station.inbound_wait_links_list]
            station.inbound_wait_links_list = [int (links) for links in station.inbound_wait_links_list]
            station.wait_links_list = list(set(station.outbound_wait_links_list+station.inbound_wait_links_list))
            g_station_list.append(station)
            g_number_of_stations += 1
        #except:
           # print('Read Error. Check your station file.')
    print('station number (+virtual station):{}'.format(g_number_of_stations))
    
    #read node data
    book = xlrd.open_workbook("input_node.xlsx")
    sheet = book.sheet_by_index(0)
    node = Node()
    node.node_id = 0
    g_node_list.append(node)
    g_number_of_nodes += 1
    for l in range(1,sheet.nrows):#read each line
        #try:
            node = Node()
            node.node_id = int(sheet.cell_value(l, 0))
            node.node_type = int(sheet.cell_value(l, 1))
            node.in_station_number = int(sheet.cell_value(l, 2))
            node.min_track_dwell_tm = int(sheet.cell_value(l,3))
            if node.node_type==1:
                all_receive_nodes_of_all_stations.append(node.node_id)
            if node.node_type==2:#receive
                all_receive_nodes_of_all_stations.append(node.node_id)
            if node.node_type==3:#depart
                all_depart_nodes_of_all_stations.append(node.node_id)
            if node.node_type==4:#depart
                all_depart_nodes_of_all_stations.append(node.node_id)
                
            g_node_list.append(node)
            g_number_of_nodes += 1
            if g_number_of_nodes % 100 == 0:
                print('reading {} nodes..' \
                      .format(g_number_of_nodes))
        #except:
            #print('Read Error. Check your node file.')
    print('node number:{}'.format(g_number_of_nodes))
    #read receive depart node run tm
    for station_i in range(1,g_number_of_stations):
        g_node_list[g_station_list[station_i].inbound_a_node].station_route_run_tm=g_station_list[station_i].station_route_run_tm
        g_node_list[g_station_list[station_i].inbound_d_node].station_route_run_tm=g_station_list[station_i].station_route_run_tm
        g_node_list[g_station_list[station_i].outbound_a_node].station_route_run_tm=g_station_list[station_i].station_route_run_tm
        g_node_list[g_station_list[station_i].outbound_d_node].station_route_run_tm=g_station_list[station_i].station_route_run_tm
    
    #read link data
    book1 = xlrd.open_workbook("input_link.xlsx")
    book2 = xlrd.open_workbook("input_conflict_station_route.xlsx")
    sheet1 = book1.sheet_by_index(0)
    sheet2 = book2.sheet_by_index(0)
    link = Link()
    link.link_id = 0
    g_link_list.append(link)
    g_number_of_links += 1
    for l in range(1,sheet1.nrows):
        try:
            link = Link()
            link.link_id = int(sheet1.cell_value(l,0))
            link.link_type = int(sheet1.cell_value(l,1))
            link.from_node_id = int(sheet1.cell_value(l,2))
            link.to_node_id = int(sheet1.cell_value(l,3))
            link.train_run_tm = int(sheet1.cell_value(l,4))
            link.link_capacity = int(sheet1.cell_value(l,5))
            link.link_cost = int(sheet1.cell_value(l,6))
            g_node_list[link.from_node_id].outgoing_node_list.append(link.to_node_id)
            g_node_list[link.from_node_id].outgoing_link_list.append(link.link_id)
            g_node_list[link.to_node_id].ingoing_link_list.append(link.link_id)
            if (link.link_type==3) or (link.link_type==4):
                link.conflict_link_list.append(link.link_id)
            g_link_list.append(link)
            g_number_of_links += 1
            if g_number_of_links % 1000 == 0:
                print('reading {} links..' \
                      .format(g_number_of_links))
        except:
            print('Read Error. Check your link file.')
    print('link number:{}'.format(g_number_of_links))
    node_link_map()
    for l in range(1,sheet2.nrows):
        link1_pre_n = int (sheet2.cell_value(l,0))
        link1_after_n = int (sheet2.cell_value(l,1))
        link2_pre_n = int (sheet2.cell_value(l,2))
        link2_after_n = int (sheet2.cell_value(l,3))
        link1 = int (node_link_map.loc[(node_link_map.from_node_id == link1_pre_n) & (node_link_map.to_node_id == link1_after_n),'link_id'].values[0])
        link2 = int (node_link_map.loc[(node_link_map.from_node_id == link2_pre_n) & (node_link_map.to_node_id == link2_after_n),'link_id'].values[0])
        g_link_list[link1].conflict_link_list.append(link2)

    for stop_node_ii in range(0,len(g_all_stations_stop_nodes_list)):
        stop_node =g_all_stations_stop_nodes_list[stop_node_ii]
        stop_link = int (node_link_map.loc[(node_link_map.from_node_id == stop_node) & (node_link_map.to_node_id == stop_node),'link_id'].values[0])
        g_all_stations_wait_links_list.append(stop_link)
    #read train data
    book = xlrd.open_workbook("input_train.xlsx")
    sheet = book.sheet_by_index(0)
    train = Train()
    train.train_id = 0
    g_train_list.append(train)
    g_number_of_trains += 1
    for l in range (1, sheet.nrows):
        #try:
            train = Train()
            train.train_id = int(sheet.cell_value(l,0))
            train.train_direction = int(sheet.cell_value(l,1))
            train.actual_orig_node_id = int(sheet.cell_value(l,2))
            train.actual_dest_node_id = int(sheet.cell_value(l,3))
            train.original_tm_beginning = int(sheet.cell_value(l,4))
            train.plan_dest_arrival_tm = int(sheet.cell_value(l,5))            
            station_sequence_list = str(sheet.cell_value(l,6))
            train.station_sequence_list = station_sequence_list.strip().split(';')
            train.station_sequence_list = [int (station) for station in train.station_sequence_list]
            stop_pattern_list = str(sheet.cell_value(l,7))
            train.stop_pattern_list = stop_pattern_list.strip().split(';')
            train.stop_pattern_list = [int (stop) for stop in train.stop_pattern_list]
            plan_station_d_tm_list = str(sheet.cell_value(l,8))
            train.plan_station_d_tm_list = plan_station_d_tm_list.strip().split(';')
            train.plan_station_d_tm_list = [int (tm) for tm in train.plan_station_d_tm_list]
            min_dwell_tm_list = str(sheet.cell_value(l,9))
            train.min_dwell_tm_list = min_dwell_tm_list.strip().split(';')
            train.min_dwell_tm_list = [int (tm) for tm in train.min_dwell_tm_list]            
            max_dwell_tm_list = str(sheet.cell_value(l,10))
            train.max_dwell_tm_list = max_dwell_tm_list.strip().split(';')
            train.max_dwell_tm_list = [int (tm) for tm in train.max_dwell_tm_list]
            track_preference = str(sheet.cell_value(l,11))
            train.track_preference = track_preference.strip().split(';')
            train.track_preference = [int (tm) for tm in train.track_preference]
            add_links_of_virtual_wt_link_at_orig_and_dest_node(train.actual_orig_node_id,0)
            add_links_of_virtual_wt_link_at_orig_and_dest_node(train.actual_dest_node_id,1)
            train.available_nodes_list = train_available_nodes_list(train.train_direction,train.stop_pattern_list,train.station_sequence_list)
            train.available_links_list = train_available_links_list(train.train_direction,train.stop_pattern_list,train.station_sequence_list,train.actual_orig_node_id,train.actual_dest_node_id)

            g_train_list.append(train)
            g_number_of_trains += 1   
            if g_number_of_trains % 50 == 0:
                print('reading {} trains..' \
                      .format(g_number_of_trains))               
        #except:                 
           # print('Read Error. Check your train file.')

    for k in range(1,g_number_of_trains):
        g_train_list[k].available_arcs_list=train_available_arcs_list(g_train_list[k].available_links_list,
                                                                      g_train_list[k].actual_orig_node_id,
                                                                      g_train_list[k].actual_dest_node_id,
                                                                      g_train_list[k].original_tm_beginning,
                                                                      g_train_list[k].plan_station_d_tm_list,
                                                                      g_train_list[k].plan_dest_arrival_tm)
           
    #read blockage data
    book = xlrd.open_workbook("input_blockage.xlsx")  # open file
    sheet = book.sheet_by_index(0)  # open sheet
    blockage = Blockage()
    blockage.blockage_id = 0
    g_blockage_list.append(blockage)
    g_number_of_blockages += 1
    for row in range(1, sheet.nrows):
        try:
            blockage = Blockage()
            blockage.blockage_id = int(sheet.cell_value(row, 0))
            blockage.blockage_node_id = int(sheet.cell_value(row, 1))
            blockage_link_list = str(sheet.cell_value(row, 2))
            blockage.blockage_link_list = blockage_link_list.strip().split(';')
            blockage.blockage_link_list = [int (link) for link in blockage.blockage_link_list]            
            blockage.from_time_interval = int(sheet.cell_value(row, 3))
            blockage.to_time_interval = int(sheet.cell_value(row, 4))
            blockage.blockage_type = int(sheet.cell_value(row, 5))
            g_blockage_list.append(blockage)
            g_number_of_blockages += 1
            if g_number_of_blockages % 50 == 0:
                print('reading {} blockages..'.format(g_number_of_blockages))
        except:
            print('Read error. Check your blockage file')
    print('blockages_number:{}'.format(g_number_of_blockages))
    
    for l in range(1,g_number_of_links):
        if g_link_list[l].link_type==4:
            calculate_conflict_link_id_of_section_link(l)


def calculate_conflict_link_id_of_section_link(link_id):
    link_id_from_node=g_link_list[link_id].from_node_id
    link_id_to_node=g_link_list[link_id].to_node_id
    for ll in range(1,g_number_of_links):
        if g_link_list[ll].link_type==4 and ll !=link_id:
            if g_link_list[ll].from_node_id==link_id_from_node and g_link_list[ll].to_node_id==link_id_to_node:
                g_link_list[link_id].conflict_link_list.append(ll)
                


def add_links_of_virtual_wt_link_at_orig_and_dest_node(node_id,orig_dest_flag):#orig_dest_flag=0 orig,=1,dest
    global all_virtual_wt_node
    global g_number_of_links
    global node_link_map
    if not node_id in all_virtual_wt_node:
        link=Link()
        link.link_id=g_number_of_links
        link.link_type=orig_dest_flag+virtual_link_basic_number
        link.from_node_id=node_id
        link.to_node_id=node_id
        link.train_run_tm=1
        link.link_capacity=1000000
        if orig_dest_flag==0:
            link.link_cost=virtual_ori_wt_arc_cost
        else:
            link.link_cost=virtual_dest_wt_arc_cost
        g_link_list.append(link)
        node_link_map = node_link_map.append({'link_id':link.link_id,\
                                              'from_node_id':node_id,\
                                              'to_node_id':node_id},ignore_index=True)        
        g_number_of_links += 1
        all_virtual_wt_node.append(node_id)
        g_node_list[node_id].outgoing_node_list.append(node_id)
        g_node_list[node_id].outgoing_link_list.append(link.link_id)
        g_node_list[node_id].ingoing_link_list.append(link.link_id)

def g_initialize_multiplier():
    # initial train link multipliers
    for l in range(1, g_number_of_links):
        g_link_list[l].time_dependent_LR_multiplier = [initial_multiplier for t in range(0, g_number_of_time_intervals + 1)]
    for i in range(1,g_number_of_nodes):
        if g_node_list[i].node_type==1 or g_node_list[i].node_type==2 or g_node_list[i].node_type==3 or g_node_list[i].node_type==4:
            g_node_list[i].time_dependent_LR_multiplier= [initial_multiplier for t in range(0, g_number_of_time_intervals + 1)]
def g_initialize_cost():
    # initial train link cost
    for l in range(1, g_number_of_links):
        g_link_list[l].time_dependent_LR_cost = [0 for t in range(0, g_number_of_time_intervals + 1)]
        g_link_list[l].time_dependent_ADMM_cost = [0 for t in range(0, g_number_of_time_intervals + 1)]    


def time_dependent_link_blockage():#float
    global time_dependent_link_blockage
    time_dependent_link_blockage = np.ones([g_number_of_links, g_number_of_time_intervals]).astype(int)
    for b in range(1, g_number_of_blockages):#t_link is 1 without block; is -1 with block； is -2 window just over,train cannot run due to the speed difference
        if g_blockage_list[b].blockage_type==0:
            for i in range(0,len(g_blockage_list[b].blockage_link_list)):
                blockage_link_id = g_blockage_list[b].blockage_link_list[i]
                link_running_time=g_link_list[blockage_link_id].train_run_tm
                from_time_interval = g_blockage_list[b].from_time_interval - link_running_time+1
                to_time_interval = g_blockage_list[b].to_time_interval
                for tt in range(from_time_interval, to_time_interval+1):
                    time_dependent_link_blockage[blockage_link_id][tt] = int(-1)
    return()


def train_time_dependent_link_cost(train_id):
    virtual_source_penalty_of_each_min_deviation=virtual_ori_wt_arc_cost
    time_dependent_link_cost = 100000 * np.ones([g_number_of_links, g_number_of_time_intervals]).astype(int)
    train_actual_orig_node_id=g_train_list[train_id].actual_orig_node_id
    train_actual_orig_node_tm=g_train_list[train_id].original_tm_beginning
    for l in range(1,g_number_of_links):
        for t in range(1,g_number_of_time_intervals):
            cost = g_link_list[l].link_cost
            time_dependent_link_cost[l][t] = int(cost)
    for first_receive_link_id in g_node_list[train_actual_orig_node_id].outgoing_link_list:
        if g_link_list[first_receive_link_id].from_node_id!=g_link_list[first_receive_link_id].to_node_id:
            for t in range(max(1,train_actual_orig_node_tm-T_orig_eariler),\
                        train_actual_orig_node_tm):
                deviation_tm_abs=abs(t-train_actual_orig_node_tm)
                time_dependent_link_cost[first_receive_link_id][t]+=(2*deviation_tm_abs-T_orig_eariler)*virtual_source_penalty_of_each_min_deviation
    for station_ii in range(0,len(g_train_list[train_id].track_preference)):
        track_preference=g_train_list[train_id].track_preference[station_ii]
        if track_preference in g_all_stations_stop_nodes_list:
            track_waiting_link=int (node_link_map.loc[(node_link_map.from_node_id == track_preference) & \
                                   (node_link_map.to_node_id == track_preference),'link_id'].values[0])
            track_preference_link_cost = g_link_list[track_waiting_link].link_cost
            for t in range(1,g_number_of_time_intervals):
                time_dependent_link_cost[track_waiting_link][t]=track_preference_link_cost-1
    g_train_list[train_id].time_dependent_link_cost= copy.deepcopy(time_dependent_link_cost)


class CVSState:
    def __init__(self):
        self.current_node_id = 0
        self.m_visit_node_sequence=[]
        self.m_visit_time_sequence=[]
        self.m_visit_link_sequence=[]
        
        self.Label_waiting_time= 0        
        self.Label_cost= MAX_LABEL_COST  #with LR and rho
        
        self.LabelCost_for_lr = MAX_LABEL_COST   #with LR price
        self.PrimalLabelCost = MAX_LABEL_COST    #without LR price



        
    def Caculate_cost(self,OLA_Flag,link_id,link_t,pElement,train_id):
        if OLA_Flag==0:
            self.PrimalLabelCost=pElement.PrimalLabelCost+g_train_list[train_id].time_dependent_link_cost[link_id][link_t]
        elif OLA_Flag==1:
            self.LabelCost_for_lr=pElement.LabelCost_for_lr+g_train_list[train_id].time_dependent_link_cost[link_id][link_t]+\
                g_link_list[link_id].time_dependent_LR_cost[link_t]
        elif OLA_Flag==2:
            self.Label_cost=pElement.Label_cost+g_train_list[train_id].time_dependent_link_cost[link_id][link_t]+\
                g_link_list[link_id].time_dependent_ADMM_cost[link_t]

    def Update_node_time_link_sequence(self,pElement,node,time,link):
        self.m_visit_node_sequence.extend(pElement.m_visit_node_sequence)
        self.m_visit_node_sequence.append(node)
        self.m_visit_time_sequence.extend(pElement.m_visit_time_sequence)
        self.m_visit_time_sequence.append(time)
        self.m_visit_link_sequence.extend(pElement.m_visit_link_sequence)
        self.m_visit_link_sequence.append(link)
        
    def Add_end_n_and_end_t(self,outgoing_node_id,next_t):
        self.m_visit_node_sequence.append(outgoing_node_id)
        self.m_visit_time_sequence.append(next_t)
        
    def generate_string_key(self):
        str=self.current_node_id
        return str
        

class C_time_indexed_state_vector:
    def __init__(self):
        self.current_time=0
        self.CVSStateVector=[[],[]]
        self.m_state_map=[[[],[]] for node in range(1,g_number_of_nodes+1)]
        
    def find_state_map_in_CVSStateVector_position(self,node_id,CVSS_ii):
        Position_indicator=self.m_state_map[node_id][1][CVSS_ii]
        Position_of_old=self.CVSStateVector[1].index(Position_indicator)
        return (Position_of_old)
    
    def update_state(self,new_element,OLA_Flag):
        new_node_id=new_element.current_node_id
        if len(self.CVSStateVector[0])==0:
            CVSS_number=1
        else:
            CVSS_number=max(self.CVSStateVector[1])+1
        if len(self.m_state_map[new_node_id][0])==0:
            self.CVSStateVector[0].append(new_element)
            self.CVSStateVector[1].append(CVSS_number)
            self.m_state_map[new_node_id][0].append(new_element)
            self.m_state_map[new_node_id][1].append(CVSS_number)
        else:
            if OLA_Flag==0:
                if not (new_node_id in g_all_stations_stop_nodes_list):
                    if new_element.PrimalLabelCost<self.m_state_map[new_node_id][0][0].PrimalLabelCost:
                        self.m_state_map[new_node_id][0].pop()
                        self.m_state_map[new_node_id][0].append(new_element)
                        Position_indicator=self.m_state_map[new_node_id][1][0]
                        Position_of_old=self.CVSStateVector[1].index(Position_indicator)
                        self.CVSStateVector[0][Position_of_old]=new_element
                elif (new_node_id in g_all_stations_stop_nodes_list):
                    Already_add_indicator=0
                    for CVSS_ii in range(0,len(self.m_state_map[new_node_id][0])):
                        if ((new_element.PrimalLabelCost>self.m_state_map[new_node_id][0][CVSS_ii].PrimalLabelCost) and \
                            (new_element.Label_waiting_time>self.m_state_map[new_node_id][0][CVSS_ii].Label_waiting_time)) or \
                            ((new_element.PrimalLabelCost<self.m_state_map[new_node_id][0][CVSS_ii].PrimalLabelCost) and \
                             (new_element.Label_waiting_time<self.m_state_map[new_node_id][0][CVSS_ii].Label_waiting_time)):
                            if Already_add_indicator==0:
                                self.CVSStateVector[0].append(new_element)
                                self.CVSStateVector[1].append(CVSS_number)
                                self.m_state_map[new_node_id][0].append(new_element)
                                self.m_state_map[new_node_id][1].append(CVSS_number)
                                Already_add_indicator+=1
                        elif ((new_element.PrimalLabelCost<self.m_state_map[new_node_id][0][CVSS_ii].PrimalLabelCost) and \
                            (new_element.Label_waiting_time==self.m_state_map[new_node_id][0][CVSS_ii].Label_waiting_time)) or \
                            ((new_element.PrimalLabelCost<self.m_state_map[new_node_id][0][CVSS_ii].PrimalLabelCost) and \
                             (new_element.Label_waiting_time>self.m_state_map[new_node_id][0][CVSS_ii].Label_waiting_time)) or\
                            ((new_element.PrimalLabelCost==self.m_state_map[new_node_id][0][CVSS_ii].PrimalLabelCost) and \
                             (new_element.Label_waiting_time>self.m_state_map[new_node_id][0][CVSS_ii].Label_waiting_time)):
                            if Already_add_indicator==0:
                                self.m_state_map[new_node_id][0].pop(CVSS_ii)
                                self.m_state_map[new_node_id][0].insert(CVSS_ii,new_element)
                                Position_indicator=self.m_state_map[new_node_id][1][CVSS_ii]
                                Position_of_old=self.CVSStateVector[1].index(Position_indicator)
                                self.CVSStateVector[0].pop(Position_of_old)
                                self.CVSStateVector[0].insert(Position_of_old,new_element)
                                Already_add_indicator+=1
                            else:
                                self.m_state_map[new_node_id][0].pop(CVSS_ii)
                                Position_indicator=self.m_state_map[new_node_id][1][CVSS_ii]
                                self.m_state_map[new_node_id][1].pop(CVSS_ii)
                                Position_of_old=self.CVSStateVector[1].index(Position_indicator)
                                self.CVSStateVector[0].pop(Position_of_old)
                                self.CVSStateVector[1].pop(Position_of_old)
            elif OLA_Flag==1:
                if not (new_node_id in g_all_stations_stop_nodes_list):
                    if new_element.LabelCost_for_lr<self.m_state_map[new_node_id][0][0].LabelCost_for_lr:
                        self.m_state_map[new_node_id][0].pop()
                        self.m_state_map[new_node_id][0].append(new_element)
                        Position_indicator=self.m_state_map[new_node_id][1][0]
                        Position_of_old=self.CVSStateVector[1].index(Position_indicator)
                        self.CVSStateVector[0].pop(Position_of_old)
                        self.CVSStateVector[0].insert(Position_of_old,new_element)
                elif (new_node_id in g_all_stations_stop_nodes_list):
                    Already_add_indicator=0
                    for CVSS_ii in range(0,len(self.m_state_map[new_node_id][0])):
                        if ((new_element.LabelCost_for_lr>self.m_state_map[new_node_id][0][CVSS_ii].LabelCost_for_lr) and \
                            (new_element.Label_waiting_time>self.m_state_map[new_node_id][0][CVSS_ii].Label_waiting_time)) or \
                            ((new_element.LabelCost_for_lr<self.m_state_map[new_node_id][0][CVSS_ii].LabelCost_for_lr) and \
                             (new_element.Label_waiting_time<self.m_state_map[new_node_id][0][CVSS_ii].Label_waiting_time)):
                            if Already_add_indicator==0:
                                self.CVSStateVector[0].append(new_element)
                                self.CVSStateVector[1].append(CVSS_number)
                                self.m_state_map[new_node_id][0].append(new_element)
                                self.m_state_map[new_node_id][1].append(CVSS_number)
                                Already_add_indicator+=1
                        elif ((new_element.LabelCost_for_lr<self.m_state_map[new_node_id][0][CVSS_ii].LabelCost_for_lr) and \
                            (new_element.Label_waiting_time==self.m_state_map[new_node_id][0][CVSS_ii].Label_waiting_time)) or \
                            ((new_element.LabelCost_for_lr<self.m_state_map[new_node_id][0][CVSS_ii].LabelCost_for_lr) and \
                             (new_element.Label_waiting_time>self.m_state_map[new_node_id][0][CVSS_ii].Label_waiting_time)) or\
                            ((new_element.LabelCost_for_lr==self.m_state_map[new_node_id][0][CVSS_ii].LabelCost_for_lr) and \
                             (new_element.Label_waiting_time>self.m_state_map[new_node_id][0][CVSS_ii].Label_waiting_time)):
                            if Already_add_indicator==0:
                                self.m_state_map[new_node_id][0].pop(CVSS_ii)
                                self.m_state_map[new_node_id][0].insert(CVSS_ii,new_element)
                                Position_indicator=self.m_state_map[new_node_id][1][CVSS_ii]
                                Position_of_old=self.CVSStateVector[1].index(Position_indicator)
                                self.CVSStateVector[0].pop(Position_of_old)
                                self.CVSStateVector[0].insert(Position_of_old,new_element)
                                Already_add_indicator+=1
                            else:
                                self.m_state_map[new_node_id][0].pop(CVSS_ii)
                                Position_indicator=self.m_state_map[new_node_id][1][CVSS_ii]
                                self.m_state_map[new_node_id][1].pop(CVSS_ii)
                                Position_of_old=self.CVSStateVector[1].index(Position_indicator)
                                self.CVSStateVector[0].pop(Position_of_old)
                                self.CVSStateVector[1].pop(Position_of_old)
            elif OLA_Flag==2:
                if not (new_node_id in g_all_stations_stop_nodes_list):
                    if new_element.Label_cost<self.m_state_map[new_node_id][0][0].Label_cost:
                        self.m_state_map[new_node_id][0].pop()
                        self.m_state_map[new_node_id][0].append(new_element)
                        Position_indicator=self.m_state_map[new_node_id][1][0]
                        Position_of_old=self.CVSStateVector[1].index(Position_indicator)
                        self.CVSStateVector[0].pop(Position_of_old)
                        self.CVSStateVector[0].insert(Position_of_old,new_element)
                elif (new_node_id in g_all_stations_stop_nodes_list):
                    Already_add_indicator=0
                    for CVSS_ii in range(0,len(self.m_state_map[new_node_id][0])):
                        if ((new_element.Label_cost<self.m_state_map[new_node_id][0][CVSS_ii].Label_cost) and \
                            (new_element.Label_waiting_time==self.m_state_map[new_node_id][0][CVSS_ii].Label_waiting_time)) or \
                            ((new_element.Label_cost<self.m_state_map[new_node_id][0][CVSS_ii].Label_cost) and \
                             (new_element.Label_waiting_time>self.m_state_map[new_node_id][0][CVSS_ii].Label_waiting_time)) or\
                            ((new_element.Label_cost==self.m_state_map[new_node_id][0][CVSS_ii].Label_cost) and \
                             (new_element.Label_waiting_time>self.m_state_map[new_node_id][0][CVSS_ii].Label_waiting_time)):
                            if Already_add_indicator==0:
                                self.m_state_map[new_node_id][0].pop(CVSS_ii)
                                self.m_state_map[new_node_id][0].insert(CVSS_ii,new_element)
                                Position_indicator=self.m_state_map[new_node_id][1][CVSS_ii]
                                Position_of_old=self.CVSStateVector[1].index(Position_indicator)
                                self.CVSStateVector[0].pop(Position_of_old)
                                self.CVSStateVector[0].insert(Position_of_old,new_element)
                                Already_add_indicator+=1
                            else:
                                self.m_state_map[new_node_id][0].pop(CVSS_ii)
                                Position_indicator=self.m_state_map[new_node_id][1][CVSS_ii]
                                Position_of_old=self.CVSStateVector[1].index(Position_indicator)
                                self.m_state_map[new_node_id][1].pop(CVSS_ii)
                                self.CVSStateVector[0].pop(Position_of_old)
                                self.CVSStateVector[1].pop(Position_of_old)
                        elif ((new_element.Label_cost>self.m_state_map[new_node_id][0][CVSS_ii].Label_cost) and \
                            (new_element.Label_waiting_time>self.m_state_map[new_node_id][0][CVSS_ii].Label_waiting_time)) or \
                            ((new_element.Label_cost<self.m_state_map[new_node_id][0][CVSS_ii].Label_cost) and \
                             (new_element.Label_waiting_time<self.m_state_map[new_node_id][0][CVSS_ii].Label_waiting_time)):
                            if Already_add_indicator==0:
                                self.CVSStateVector[0].append(new_element)
                                self.CVSStateVector[1].append(CVSS_number)
                                self.m_state_map[new_node_id][0].append(new_element)
                                self.m_state_map[new_node_id][1].append(CVSS_number)
                                Already_add_indicator+=1            
                
    def Reset(self):
        self.current_time = 0
        self.CVSStateVector=[[],[]]
        self.m_state_map=[[[],[]] for node in range(1,g_number_of_nodes+1)]
        
    def Sort(self,OLA_Flag):
        self.CVSStateVector[0].sort(key=lambda x:x.Label_waiting_time,reverse=True)
        if OLA_Flag ==0: #primal_cost/upper_bound
            self.CVSStateVector[0]=sorted(self.CVSStateVector[0],key=lambda x:x.PrimalLabelCost)
        if OLA_Flag ==1: #LR
            self.CVSStateVector[0] = sorted(self.CVSStateVector[0],key=lambda x:x.LabelCost_for_lr)
        if OLA_Flag ==2: #ADMM
            self.CVSStateVector[0] = sorted(self.CVSStateVector[0],key=lambda x:x.Label_cost)

def g_time_dependent_dynamic_programming_for_trains():
    # dynamic programming for each train
    global time_dependent_link_volume_for_trains
    global time_dependent_implied_waiting_link_flag_for_trains
    global C_train_time_dependent_state_vector
    global C_train_end_state_vector
    
    C_train_time_dependent_state_vector = [[None for t in range(1,g_number_of_time_intervals+1)] for k in range(1,g_number_of_trains+1)]
    C_train_end_state_vector = [None for k in range(1,g_number_of_trains+1)]
    for k in range(1, g_number_of_trains):
        # get agent information

        current_train_orig_node_id = g_train_list[k].actual_orig_node_id
        current_train_dest_node_id = g_train_list[k].actual_dest_node_id
        original_tm_beginning = int(g_train_list[k].original_tm_beginning)
        #plan_dest_arrival_tm = int(g_train_list[k].plan_dest_arrival_tm)
        volume = 1
        #delete current link_volumn for all_trains indicator
        if len(g_train_list[k].link_sequence)>0:
            for l in range(0, len(g_train_list[k].link_sequence)):
                link_id = g_train_list[k].link_sequence[l]
                time = g_train_list[k].time_sequence[l]
                time_dependent_link_volume_for_trains[link_id][time] = time_dependent_link_volume_for_trains[link_id][
                                                                          time] - volume
                
        for l in g_all_stations_wait_links_list:
            for t in range(1,g_number_of_time_intervals):
                if g_train_list[k].time_dependent_implied_waiting_link_flag[l][t]!=0:
                    time_dependent_implied_waiting_link_flag_for_trains[l][t] = time_dependent_implied_waiting_link_flag_for_trains[l][t]- volume

                    
        #reset the parameter of current train
        g_train_list[k].time_dependent_link_volume=[[0 for t in range(1, g_number_of_time_intervals + 1)] for l in
                                                      range(1, g_number_of_links + 1)]
        g_train_list[k].time_dependent_implied_waiting_link_flag = [[0 for t in range(1, g_number_of_time_intervals + 1)] for l in
                                                      range(1, g_number_of_links + 1)]

        #calculate ADMM without current train

        calculate_LR_and_ADMM_for_current_train_1(time_dependent_link_volume_for_trains,time_dependent_implied_waiting_link_flag_for_trains,k)


        #reset
        g_train_list[k].node_sequence = []
        g_train_list[k].time_sequence = []
        g_train_list[k].link_sequence = []


        # dynamic programming
        g_train_list[k].Label_cost=0

        for t in range(0,g_number_of_time_intervals):
            C_train_time_dependent_state_vector[k][t] = C_time_indexed_state_vector()
            C_train_time_dependent_state_vector[k][t].Reset()
            C_train_time_dependent_state_vector[k][t].current_time=t
    

        C_train_end_state_vector[k]= C_time_indexed_state_vector()
        C_train_end_state_vector[k].Reset()
        #original_node
        element=CVSState()
        element.current_node_id=current_train_orig_node_id
        element.Label_cost=0
        element.PrimalLabelCost=0
        C_train_time_dependent_state_vector[k][original_tm_beginning].update_state(element,2)


        for t in range(original_tm_beginning, g_number_of_time_intervals):
            C_train_time_dependent_state_vector[k][t].Sort(2)
            for w_index in range(min(BestKSize, len(C_train_time_dependent_state_vector[k][t].CVSStateVector[0]))):
                pElement=C_train_time_dependent_state_vector[k][t].CVSStateVector[0][w_index]
                n=pElement.current_node_id
                if n==current_train_dest_node_id:
                    continue
                for outgoing_link_id in g_node_list[n].outgoing_link_list:            
                    if (outgoing_link_id in g_train_list[k].available_links_list):
                        outgoing_node_id=g_link_list[outgoing_link_id].to_node_id
                        train_run_tm = g_link_list[outgoing_link_id].train_run_tm
                        if (time_dependent_link_blockage[outgoing_link_id][t] == 1) and g_train_list[k].available_arcs_list[outgoing_link_id][t]==1:
                            next_t=t+train_run_tm
                            if (next_t<g_number_of_time_intervals):
                                if (n in g_all_stations_stop_nodes_list) and (not (outgoing_node_id in g_all_stations_stop_nodes_list)):
                                    sta_ii=get_sta_ii_of_waiting_node_train_id_ranked(n,k)                                        
                                    stop_pattern = g_train_list[k].stop_pattern_list[sta_ii]
                                    max_track_dwell_time=g_train_list[k].max_dwell_tm_list[sta_ii]
                                    if stop_pattern==1:                                   
                                        t_leave = g_train_list[k].plan_station_d_tm_list[sta_ii]
                                        min_train_dwell_time = g_train_list[k].min_dwell_tm_list[sta_ii]
                                        min_track_dwell_time = g_node_list[n].min_track_dwell_tm
                                        if (t>=t_leave) and pElement.Label_waiting_time>=min_train_dwell_time and \
                                            pElement.Label_waiting_time>=min_track_dwell_time and pElement.Label_waiting_time<=max_track_dwell_time:
                                            new_element=CVSState()
                                            new_element.current_node_id=outgoing_node_id
                                            new_element.Label_waiting_time=0
                                            new_element.Caculate_cost(2,outgoing_link_id,t,pElement,k)
                                            new_element.Caculate_cost(0,outgoing_link_id,t,pElement,k)
                                            waiting_link_id= int (node_link_map.loc[(node_link_map.from_node_id == n) & \
                                                       (node_link_map.to_node_id == n),'link_id'].values[0])
                                            for tt in range(t,min(t+headway_track,g_number_of_time_intervals)):
                                                new_element.Label_cost+=g_link_list[waiting_link_id].time_dependent_ADMM_cost[tt]
                                            new_element.Update_node_time_link_sequence(pElement,n,t,outgoing_link_id)
                                            C_train_time_dependent_state_vector[k][next_t].update_state(new_element,2) 
                                            if new_element.current_node_id==current_train_dest_node_id:
                                                new_element.Add_end_n_and_end_t(current_train_dest_node_id,next_t)
                                                C_train_end_state_vector[k].CVSStateVector[0].append(new_element) 
                                    else:
                                        min_track_dwell_time = g_node_list[n].min_track_dwell_tm
                                        if pElement.Label_waiting_time>=min_track_dwell_time and pElement.Label_waiting_time<=max_track_dwell_time:
                                            new_element=CVSState()
                                            new_element.current_node_id=outgoing_node_id
                                            new_element.Label_waiting_time=0
                                            new_element.Caculate_cost(2,outgoing_link_id,t,pElement,k)
                                            new_element.Caculate_cost(0,outgoing_link_id,t,pElement,k)
                                            waiting_link_id= int (node_link_map.loc[(node_link_map.from_node_id == n) & \
                                                       (node_link_map.to_node_id == n),'link_id'].values[0])
                                            for tt in range(t,min(t+headway_track,g_number_of_time_intervals)):
                                                new_element.Label_cost+=g_link_list[waiting_link_id].time_dependent_ADMM_cost[tt]
                                            new_element.Update_node_time_link_sequence(pElement,n,t,outgoing_link_id)
                                            C_train_time_dependent_state_vector[k][next_t].update_state(new_element,2)
                                            if new_element.current_node_id==current_train_dest_node_id:
                                                new_element.Add_end_n_and_end_t(current_train_dest_node_id,next_t)
                                                C_train_end_state_vector[k].CVSStateVector[0].append(new_element) 
                                elif (n==outgoing_node_id) and n!=current_train_orig_node_id:
                                    new_element=CVSState()
                                    new_element.current_node_id=outgoing_node_id
                                    new_element.Label_waiting_time=pElement.Label_waiting_time+1
                                    sta_ii=get_sta_ii_of_waiting_node_train_id_ranked(n,k)  
                                    if new_element.Label_waiting_time<=g_train_list[k].max_dwell_tm_list[sta_ii]:
                                        new_element.Caculate_cost(2,outgoing_link_id,t,pElement,k)
                                        new_element.Caculate_cost(0,outgoing_link_id,t,pElement,k)
                                        new_element.Update_node_time_link_sequence(pElement,n,t,outgoing_link_id)
                                        C_train_time_dependent_state_vector[k][next_t].update_state(new_element,2)
                                else:
                                    new_element=CVSState()
                                    new_element.current_node_id=outgoing_node_id
                                    new_element.Label_waiting_time=pElement.Label_waiting_time
                                    new_element.Caculate_cost(2,outgoing_link_id,t,pElement,k)
                                    new_element.Caculate_cost(0,outgoing_link_id,t,pElement,k)
                                    new_element.Update_node_time_link_sequence(pElement,n,t,outgoing_link_id)
                                    C_train_time_dependent_state_vector[k][next_t].update_state(new_element,2)
                                    if new_element.current_node_id==current_train_dest_node_id:
                                        new_element.Add_end_n_and_end_t(current_train_dest_node_id,next_t)
                                        C_train_end_state_vector[k].CVSStateVector[0].append(new_element)

        if len(C_train_end_state_vector[k].CVSStateVector[0])==0:
            g_train_list[k].Label_cost=MAX_LABEL_COST
            print('can not find space-time path for train:{}'.format(k))
            g_train_list[k].current_iteration_upper_bound_priority=MAX_LABEL_COST/g_train_list[k].train_free_flow_time
        else:
            C_train_end_state_vector[k].CVSStateVector[0].sort(key=lambda x:x.Label_cost)
            g_train_list[k].node_sequence=C_train_end_state_vector[k].CVSStateVector[0][0].m_visit_node_sequence
            g_train_list[k].time_sequence=C_train_end_state_vector[k].CVSStateVector[0][0].m_visit_time_sequence
            g_train_list[k].link_sequence=C_train_end_state_vector[k].CVSStateVector[0][0].m_visit_link_sequence
            g_train_list[k].Label_cost=C_train_end_state_vector[k].CVSStateVector[0][0].Label_cost
            g_train_list[k].PrimalLabelCost=C_train_end_state_vector[k].CVSStateVector[0][0].PrimalLabelCost
            C_train_end_state_vector[k].Reset()
            for link_i in range(0,len(g_train_list[k].link_sequence)):
                l=g_train_list[k].link_sequence[link_i]
                t=g_train_list[k].time_sequence[link_i]
                time_dependent_link_volume_for_trains[l][t] += volume
                
                g_train_list[k].time_dependent_link_volume[l][t] += volume
                if g_link_list[l].link_type==1:#departure
                    link_waiting_node=g_link_list[l].from_node_id
                    if link_waiting_node in g_all_stations_stop_nodes_list:
                        stop_link_id = int (node_link_map.loc[(node_link_map.from_node_id == link_waiting_node) & \
                                                       (node_link_map.to_node_id == link_waiting_node),'link_id'].values[0])
                    # update time-dependent link volume
                        for ttt in range(0,headway_track):
                            g_train_list[k].time_dependent_implied_waiting_link_flag[stop_link_id][t+ttt]+=volume
                            time_dependent_implied_waiting_link_flag_for_trains[stop_link_id][t+ttt]+=volume   
                            
            g_train_list[k].current_iteration_upper_bound_priority=(g_train_list[k].time_sequence[-1]-\
                                                               g_train_list[k].time_sequence[0])/g_train_list[k].train_free_flow_time
            train_result[iteration_step]+=g_train_list[k].PrimalLabelCost

        


    
    #calculate each train conflict trains number for the priority in the upper bound
    for k in range(1,g_number_of_trains):
        g_train_list[k].conflict_trains_list = []

    for l in range(1, g_number_of_links):
        if (g_link_list[l].link_type == 1) or (g_link_list[l].link_type == 2) or (g_link_list[l].link_type == 4) :#rece or depart route in station
            for k in range(1, g_number_of_trains):         
                if g_link_list[l].link_id in g_train_list[k].available_links_list :
                    for t in range(1, g_number_of_time_intervals):
                        if g_train_list[k].time_dependent_link_volume[l][t] == 1:
                            for i in range(0,len(g_link_list[l].IN_after_arcs_of_td_link[t])):
                                ll=g_link_list[l].IN_after_arcs_of_td_link[t][i][0]
                                tt=g_link_list[l].IN_after_arcs_of_td_link[t][i][1]
                                for kk in range(1, g_number_of_trains):
                                    if kk != k:
                                        if g_train_list[kk].time_dependent_link_volume[ll][tt] == 1:
                                            g_train_list[k].conflict_trains_list.append(kk)
                                            g_train_list[k].conflict_trains_list=list(set(g_train_list[k].conflict_trains_list))
                                            g_train_list[kk].conflict_trains_list.append(k)
                                            g_train_list[kk].conflict_trains_list=list(set(g_train_list[kk].conflict_trains_list))
        elif (g_link_list[l].link_type == 3) :#wait arc
            for k in range(1, g_number_of_trains):
                if g_link_list[l].link_id in g_train_list[k].available_links_list :
                    for t in range(1, g_number_of_time_intervals):
                        if (g_train_list[k].time_dependent_link_volume[l][t] == 1) or \
                            (g_train_list[k].time_dependent_implied_waiting_link_flag[l][t] == 1):
                            for kk in range(k, g_number_of_trains):
                                if kk != k:
                                    if (g_train_list[kk].time_dependent_link_volume[l][t] == 1) or \
                                        (g_train_list[kk].time_dependent_implied_waiting_link_flag[l][t] == 1):
                                        g_train_list[k].conflict_trains_list.append(kk)
                                        g_train_list[k].conflict_trains_list=list(set(g_train_list[k].conflict_trains_list))
                                        g_train_list[kk].conflict_trains_list.append(k)
                                        g_train_list[kk].conflict_trains_list=list(set(g_train_list[kk].conflict_trains_list))  

    result_feasibility[iteration_step]=1
    for k in range(1,g_number_of_trains):
        if len(g_train_list[k].conflict_trains_list)>0:
            result_feasibility[iteration_step]=0
            break
        
        
        

def check_train_result_feasibility():
    feasibility=1
    for l in range(1,g_number_of_links):
        if g_link_list[l].link_type>virtual_link_basic_number:
            break
        if feasibility==0:
            break
        for t in range(1, g_number_of_time_intervals):
            if (g_link_list[l].link_type == 1) or (g_link_list[l].link_type == 2) or (g_link_list[l].link_type == 4):                
                sum_of_trains_1_2_4=0
                for i in range(0,len(g_link_list[l].IN_after_arcs_of_td_link[t])):                    
                    ll=g_link_list[l].IN_after_arcs_of_td_link[t][i][0]
                    tt=g_link_list[l].IN_after_arcs_of_td_link[t][i][1]
                    sum_of_trains_1_2_4+=time_dependent_link_volume_for_trains[ll][tt]
                if (sum_of_trains_1_2_4<=1):
                    continue
                else:
                    feasibility=0
                    break
            if (g_link_list[l].link_type == 3):# wating arc
                if (time_dependent_link_volume_for_trains[l][t]+time_dependent_implied_waiting_link_flag_for_trains[l][t]<=1):
                    continue
                else:
                    feasibility=0
                    break
    return(feasibility)
         

def calculate_LR_for_current_iteration_4():
    
    for l in range(1, g_number_of_links):
        g_link_list[l].time_dependent_LR_cost = [0 for t in range(1, g_number_of_time_intervals + 1)]
    
        
    for i in all_receive_nodes_of_all_stations:
        if iteration_step==0 and initial_multiplier==0:
            break
        receive_route_run_tm_of_i=g_node_list[i].station_route_run_tm
        for l in g_node_list[i].outgoing_link_list:
            if g_link_list[l].link_type<virtual_link_basic_number:
                for t in range(1,g_number_of_time_intervals):
                    for tt in range(max(1,t-receive_route_run_tm_of_i-headway_route+1),t+1):
                        g_link_list[l].time_dependent_LR_cost[t]+=g_node_list[i].time_dependent_LR_multiplier[tt]
    for j in all_depart_nodes_of_all_stations:
        if iteration_step==0 and initial_multiplier==0:
            break
        depart_route_run_tm_of_j=g_node_list[j].station_route_run_tm
        for l in g_node_list[j].ingoing_link_list:
            if g_link_list[l].link_type<virtual_link_basic_number:  
                for tuo in range(1,g_number_of_time_intervals):
                    t=tuo-depart_route_run_tm_of_j
                    if t>=1:
                        for tuotuo in range(max(1,tuo-depart_route_run_tm_of_j-headway_route+1),tuo+1):
                            g_link_list[l].time_dependent_LR_cost[t]+=g_node_list[j].time_dependent_LR_multiplier[tuotuo]
                
    for l in range (1,g_number_of_links):
        if iteration_step==0 and initial_multiplier==0:
            break
        if g_link_list[l].link_type>virtual_link_basic_number:
            break
        for t in range(1, g_number_of_time_intervals):
            if (g_link_list[l].link_type == 2):#receive 
                run_tm_of_l=g_link_list[l].train_run_tm
                for tt in range(max(1,t-run_tm_of_l-headway_route+1),t+1):
                    g_link_list[l].time_dependent_LR_cost[t]+=g_link_list[l].time_dependent_LR_multiplier[tt]
            if (g_link_list[l].link_type == 1):#depart
                for i in range(0,len(g_link_list[l].IN_before_arcs_of_td_link[t])):                    
                    ll=g_link_list[l].IN_before_arcs_of_td_link[t][i][0]
                    tt=g_link_list[l].IN_before_arcs_of_td_link[t][i][1]
                    if g_link_list[ll].link_type!=1:
                        g_link_list[l].time_dependent_LR_cost[t]+=  g_link_list[ll].time_dependent_LR_multiplier[tt]
            if (g_link_list[l].link_type == 4):#section
                for i in range(0,len(g_link_list[l].IN_before_arcs_of_td_link[t])):                    
                    ll=g_link_list[l].IN_before_arcs_of_td_link[t][i][0]
                    tt=g_link_list[l].IN_before_arcs_of_td_link[t][i][1]
                    g_link_list[l].time_dependent_LR_cost[t]+=  g_link_list[ll].time_dependent_LR_multiplier[tt]
            if (g_link_list[l].link_type == 3):# wating arc
                g_link_list[l].time_dependent_LR_cost[t]= g_link_list[l].time_dependent_LR_multiplier[t]

def calculate_LR_and_ADMM_for_current_train(td_link_volume_for_trains_flag,td_implied_waiting_link_flag_for_trains_flag):
    #as Zhang Yong Xiang
    for l in range (1,g_number_of_links):
        if g_link_list[l].link_type>virtual_link_basic_number:
            break
        for t in range(1, g_number_of_time_intervals):
            g_link_list[l].time_dependent_ADMM_cost[t]=0
            if (g_link_list[l].link_type == 1) or (g_link_list[l].link_type == 2) or (g_link_list[l].link_type == 4):                
                for i in range(0,len(g_link_list[l].IN_before_arcs_of_td_link[t])):                    
                    ll=g_link_list[l].IN_before_arcs_of_td_link[t][i][0]
                    tt=g_link_list[l].IN_before_arcs_of_td_link[t][i][1]
                    LR_ll_tt= g_link_list[ll].time_dependent_LR_multiplier[tt]
                    sum_for_before_train_other_trains=0
                    for ii in range(0,len(g_link_list[ll].IN_after_arcs_of_td_link[tt])):
                        llll=g_link_list[ll].IN_after_arcs_of_td_link[tt][ii][0]
                        tttt=g_link_list[ll].IN_after_arcs_of_td_link[tt][ii][1]
                        sum_for_before_train_other_trains+=td_link_volume_for_trains_flag[llll][tttt]
                    ADMM_ll_tt=roh/2*(2*sum_for_before_train_other_trains-1)
                    g_link_list[l].time_dependent_ADMM_cost[t]+=max(0,LR_ll_tt+ADMM_ll_tt)
            if (g_link_list[l].link_type == 3):# wating arc
                LR_l_t= g_link_list[l].time_dependent_LR_multiplier[t]
                ADMM_l_t=u/2*(2*(td_link_volume_for_trains_flag[l][t]+td_implied_waiting_link_flag_for_trains_flag[l][t])-1)
                g_link_list[l].time_dependent_ADMM_cost[t]=max(0,LR_l_t+ADMM_l_t)

def calculate_LR_and_ADMM_for_current_train_1(td_link_volume_for_trains_flag,td_implied_waiting_link_flag_for_trains_flag,train_id):
    #as formulation
    original_tm_beginning=g_train_list[train_id].original_tm_beginning
    for l in range(1,g_number_of_links):
        g_link_list[l].time_dependent_ADMM_cost = [0 for t in range(0, g_number_of_time_intervals + 1)]
    for l in range(1, g_number_of_links):
        if g_link_list[l].link_id in g_train_list[train_id].available_links_list:
            for t in range(max(1,original_tm_beginning-T_orig_eariler), g_number_of_time_intervals):
                if (t>1) and (g_train_list[train_id].available_arcs_list[l][t]-g_train_list[train_id].available_arcs_list[l][t-1]==-2):
                    break                    
                if g_train_list[train_id].available_arcs_list[l][t]==1:
                    g_link_list[l].time_dependent_ADMM_cost[t] = g_link_list[l].time_dependent_LR_cost[t]
    for i in all_receive_nodes_of_all_stations:
        receive_route_run_tm_of_i=g_node_list[i].station_route_run_tm
        for l in g_node_list[i].outgoing_link_list:
            if l in g_train_list[train_id].available_links_list:
                if g_link_list[l].link_type<virtual_link_basic_number:
                    for t in range(max(1,original_tm_beginning-T_orig_eariler),g_number_of_time_intervals):
                        if (t>1) and (g_train_list[train_id].available_arcs_list[l][t]-g_train_list[train_id].available_arcs_list[l][t-1]==-2):
                            break
                        if g_train_list[train_id].available_arcs_list[l][t]==1:
                            for tt in range(max(1,t-receive_route_run_tm_of_i-headway_route+1),t+1):
                                sum_for_before_train_after_other_trains_receive_node=0
                                for ii in range(0,len(g_node_list[i].IN_after_arcs_of_td_receive_node[tt])):
                                    ll=g_node_list[i].IN_after_arcs_of_td_receive_node[tt][ii][0]
                                    ttt=g_node_list[i].IN_after_arcs_of_td_receive_node[tt][ii][1]
                                    sum_for_before_train_after_other_trains_receive_node+=td_link_volume_for_trains_flag[ll][ttt]
                                if sum_for_before_train_after_other_trains_receive_node!=0:
                                    g_link_list[l].time_dependent_ADMM_cost[t]+=roh1/2*(2*sum_for_before_train_after_other_trains_receive_node-1)

    for j in all_depart_nodes_of_all_stations:
        depart_route_run_tm_of_j=g_node_list[j].station_route_run_tm
        for l in g_node_list[j].ingoing_link_list:
            if l in g_train_list[train_id].available_links_list:
                if g_link_list[l].link_type<virtual_link_basic_number:  
                    for tuo in range(1,g_number_of_time_intervals):
                        t=tuo-depart_route_run_tm_of_j
                        if t>=1:
                            if (t>1) and (g_train_list[train_id].available_arcs_list[l][t]-g_train_list[train_id].available_arcs_list[l][t-1]==-2):
                                break
                            if g_train_list[train_id].available_arcs_list[l][t]==1:
                                for tuotuo in range(max(1,tuo-depart_route_run_tm_of_j-headway_route+1),tuo+1):
                                    sum_for_before_train_after_other_trains_depart_node=0
                                    for jj in range(0,len(g_node_list[j].IN_after_arcs_of_td_depart_node[tuotuo])):
                                        ll=g_node_list[j].IN_after_arcs_of_td_depart_node[tuotuo][jj][0]
                                        tt=g_node_list[j].IN_after_arcs_of_td_depart_node[tuotuo][jj][1]
                                        sum_for_before_train_after_other_trains_depart_node+=td_link_volume_for_trains_flag[ll][tt]
                                    if sum_for_before_train_after_other_trains_depart_node>0:
                                        g_link_list[l].time_dependent_ADMM_cost[t]+=roh2/2*(2*sum_for_before_train_after_other_trains_depart_node-1)
    for l in range (1,g_number_of_links):
        if l in g_train_list[train_id].available_links_list:
            if g_link_list[l].link_type>virtual_link_basic_number:
                break
            for t in range(max(1,original_tm_beginning-T_orig_eariler), g_number_of_time_intervals):
                if (g_link_list[l].link_type == 2):#receive 
                    run_tm_of_l=g_link_list[l].train_run_tm
                    for tt in range(max(1,t-run_tm_of_l-headway_route+1),t+1):
                        sum_for_after_other_trains_2=0
                        for i in range(0,len(g_link_list[l].IN_after_arcs_of_td_link[tt])):
                            ll=g_link_list[l].IN_after_arcs_of_td_link[tt][i][0]
                            if not(g_link_list[ll].from_node_id==g_link_list[l].from_node_id and \
                                   g_link_list[ll].to_node_id!=g_link_list[l].to_node_id):
                                ttt=g_link_list[l].IN_after_arcs_of_td_link[tt][i][1]
                                sum_for_after_other_trains_2+=td_link_volume_for_trains_flag[ll][ttt]
                        if sum_for_after_other_trains_2!=0:
                            g_link_list[l].time_dependent_ADMM_cost[t]+=roh3/2*(2*sum_for_after_other_trains_2-1)
                if (g_link_list[l].link_type == 1):#depart
                    for i in range(0,len(g_link_list[l].IN_before_arcs_of_td_link[t])):                    
                        ll=g_link_list[l].IN_before_arcs_of_td_link[t][i][0]
                        tt=g_link_list[l].IN_before_arcs_of_td_link[t][i][1]
                        if g_link_list[ll].link_type!=1:
                            sum_for_after_other_trains_1=0
                            for ii in range(0,len(g_link_list[ll].IN_after_arcs_of_td_link[tt])):
                                llll=g_link_list[ll].IN_after_arcs_of_td_link[tt][ii][0]
                                if not (g_link_list[llll].from_node_id==g_link_list[ll].from_node_id and \
                                    g_link_list[llll].to_node_id!=g_link_list[ll].to_node_id):
                                    tttt=g_link_list[ll].IN_after_arcs_of_td_link[tt][ii][1]
                                    sum_for_after_other_trains_1+=td_link_volume_for_trains_flag[llll][tttt]
                            if sum_for_after_other_trains_1>0:
                                g_link_list[l].time_dependent_ADMM_cost[t]+=roh3/2*(2*sum_for_after_other_trains_1-1)
                if (g_link_list[l].link_type == 4):#section              
                    for i in range(0,len(g_link_list[l].IN_before_arcs_of_td_link[t])):                    
                        ll=g_link_list[l].IN_before_arcs_of_td_link[t][i][0]
                        tt=g_link_list[l].IN_before_arcs_of_td_link[t][i][1]
                        sum_for_before_train_other_trains=0
                        for ii in range(0,len(g_link_list[ll].IN_after_arcs_of_td_link[tt])):
                            llll=g_link_list[ll].IN_after_arcs_of_td_link[tt][ii][0]
                            tttt=g_link_list[ll].IN_after_arcs_of_td_link[tt][ii][1]
                            sum_for_before_train_other_trains+=td_link_volume_for_trains_flag[llll][tttt]
                        if sum_for_before_train_other_trains>0:
                            g_link_list[l].time_dependent_ADMM_cost[t]+=roh4/2*(2*sum_for_before_train_other_trains-1)
                if (g_link_list[l].link_type == 3):# wating arc
                    g_link_list[l].time_dependent_ADMM_cost[t]+=max(0,roh5/2*(2*(td_link_volume_for_trains_flag[l][t]+td_implied_waiting_link_flag_for_trains_flag[l][t])-1))                       


def g_calcualte_train_result():#
    #calculate train_result
    for k in range(1,g_number_of_trains):
        for link_i in range(0,len(g_train_list[k].link_sequence)):
            l=g_train_list[k].link_sequence[link_i]
            if g_link_list[l].link_type>virtual_link_basic_number:
                continue
            t=g_train_list[k].time_sequence[link_i]
            if consider_waiting_arc_flag==1:
                train_result[iteration_step]+=g_train_list[k].time_dependent_link_cost[l][t]
            else:
                if g_link_list[l].link_type == 1 or g_link_list[l].link_type == 2 or \
                            g_link_list[l].link_type == 3 or g_link_list[l].link_type==4:
                    train_result[iteration_step]+=g_train_list[k].time_dependent_link_cost[l][t]
                    
    for l in range(1,g_number_of_links):
        if g_link_list[l].link_type>virtual_link_basic_number:
            break
        for t in range(1, g_number_of_time_intervals):
            if (g_link_list[l].link_type == 1) or (g_link_list[l].link_type == 2) or (g_link_list[l].link_type == 4):                
                sum_of_trains_1_2_4=0
                for i in range(0,len(g_link_list[l].IN_after_arcs_of_td_link[t])):                    
                    ll=g_link_list[l].IN_after_arcs_of_td_link[t][i][0]
                    tt=g_link_list[l].IN_after_arcs_of_td_link[t][i][1]
                    sum_of_trains_1_2_4+=time_dependent_link_volume_for_trains[ll][tt]
                train_result[iteration_step]+=g_link_list[l].time_dependent_LR_multiplier[t]*(sum_of_trains_1_2_4-1)
            if (g_link_list[l].link_type == 3):# wating arc
                train_result[iteration_step]+= g_link_list[l].time_dependent_LR_multiplier[t]*(time_dependent_link_volume_for_trains[l][t]+time_dependent_implied_waiting_link_flag_for_trains[l][t]-1)

    
def g_calcualte_lower_bound_2():#
    global C_train_time_dependent_state_vector_lower_bound
    global C_train_end_state_vector_lower_bound
    
    C_train_time_dependent_state_vector_lower_bound = [[None for t in range(1,g_number_of_time_intervals+1)] for k in range(1,g_number_of_trains+1)]
    C_train_end_state_vector_lower_bound = [None for k in range(1,g_number_of_trains+1)]

    for k in range(1, g_number_of_trains):
        # get agent information
        current_train_orig_node_id = g_train_list[k].actual_orig_node_id
        current_train_dest_node_id = g_train_list[k].actual_dest_node_id
        original_tm_beginning = int(g_train_list[k].original_tm_beginning)
        
        #reset
        g_train_list[k].node_sequence_lower_bound = []
        g_train_list[k].time_sequence_lower_bound = []
        g_train_list[k].link_sequence_lower_bound = []
        
        # dynamic programming
        for t in range(0,g_number_of_time_intervals):
            C_train_time_dependent_state_vector_lower_bound[k][t] = C_time_indexed_state_vector()
            C_train_time_dependent_state_vector_lower_bound[k][t].Reset()
            C_train_time_dependent_state_vector_lower_bound[k][t].current_time=t
    

        C_train_end_state_vector_lower_bound[k]= C_time_indexed_state_vector()
        C_train_end_state_vector_lower_bound[k].Reset()
        #original_node
        element=CVSState()
        element.current_node_id=current_train_orig_node_id
        element.LabelCost_for_lr=0
        C_train_time_dependent_state_vector_lower_bound[k][original_tm_beginning].update_state(element,1)


        for t in range(original_tm_beginning, g_number_of_time_intervals):

            C_train_time_dependent_state_vector_lower_bound[k][t].Sort(1)
            for w_index in range(min(BestKSize, len(C_train_time_dependent_state_vector_lower_bound[k][t].CVSStateVector[0]))):
                pElement=C_train_time_dependent_state_vector_lower_bound[k][t].CVSStateVector[0][w_index]
                n=pElement.current_node_id
                if n==current_train_dest_node_id:
                    continue
                for outgoing_link_id in g_node_list[n].outgoing_link_list:            
                    if (outgoing_link_id in g_train_list[k].available_links_list):
                        outgoing_node_id=g_link_list[outgoing_link_id].to_node_id
                        train_run_tm = g_link_list[outgoing_link_id].train_run_tm
                        if (time_dependent_link_blockage[outgoing_link_id][t] == 1)  and g_train_list[k].available_arcs_list[outgoing_link_id][t]==1:
                            next_t=t+train_run_tm
                            if (next_t<g_number_of_time_intervals):
                                if (n in g_all_stations_stop_nodes_list) and (not (outgoing_node_id in g_all_stations_stop_nodes_list)):
                                    sta_ii=get_sta_ii_of_waiting_node_train_id_ranked(n,k)                                        
                                    stop_pattern = g_train_list[k].stop_pattern_list[sta_ii]
                                    max_track_dwell_time=g_train_list[k].max_dwell_tm_list[sta_ii]
                                    if stop_pattern==1:                                   
                                        t_leave = g_train_list[k].plan_station_d_tm_list[sta_ii]
                                        min_train_dwell_time = g_train_list[k].min_dwell_tm_list[sta_ii]
                                        min_track_dwell_time = g_node_list[n].min_track_dwell_tm
                                        if (t>=t_leave) and pElement.Label_waiting_time>=min_train_dwell_time and \
                                            pElement.Label_waiting_time>=min_track_dwell_time and pElement.Label_waiting_time<=max_track_dwell_time:# 
                                            new_element=CVSState()
                                            new_element.current_node_id=outgoing_node_id
                                            #new_element.Label_waiting_time=0
                                            new_element.Caculate_cost(1,outgoing_link_id,t,pElement,k)                                            
                                            waiting_link_id= int (node_link_map.loc[(node_link_map.from_node_id == n) & \
                                                       (node_link_map.to_node_id == n),'link_id'].values[0])
                                            for tt in range(t,min(t+headway_track,g_number_of_time_intervals)):
                                                new_element.LabelCost_for_lr+=g_link_list[waiting_link_id].time_dependent_LR_cost[tt]
                                            new_element.Update_node_time_link_sequence(pElement,n,t,outgoing_link_id)
                                            C_train_time_dependent_state_vector_lower_bound[k][next_t].update_state(new_element,1) 
                                            if new_element.current_node_id==current_train_dest_node_id:
                                                new_element.Add_end_n_and_end_t(current_train_dest_node_id,next_t)
                                                C_train_end_state_vector_lower_bound[k].CVSStateVector[0].append(new_element) 
                                    else:
                                        min_track_dwell_time = g_node_list[n].min_track_dwell_tm
                                        if pElement.Label_waiting_time>=min_track_dwell_time and pElement.Label_waiting_time<=max_track_dwell_time:
                                            new_element=CVSState()
                                            new_element.current_node_id=outgoing_node_id
                                            #new_element.Label_waiting_time=0
                                            new_element.Caculate_cost(1,outgoing_link_id,t,pElement,k)
                                            waiting_link_id= int (node_link_map.loc[(node_link_map.from_node_id == n) & \
                                                       (node_link_map.to_node_id == n),'link_id'].values[0])
                                            for tt in range(t,min(t+headway_track,g_number_of_time_intervals)):
                                                new_element.LabelCost_for_lr+=g_link_list[waiting_link_id].time_dependent_LR_cost[tt]
                                            new_element.Update_node_time_link_sequence(pElement,n,t,outgoing_link_id)
                                            C_train_time_dependent_state_vector_lower_bound[k][next_t].update_state(new_element,1)
                                            if new_element.current_node_id==current_train_dest_node_id:
                                                new_element.Add_end_n_and_end_t(current_train_dest_node_id,next_t)
                                                C_train_end_state_vector_lower_bound[k].CVSStateVector[0].append(new_element) 
                                elif (n==outgoing_node_id) and n!=current_train_orig_node_id:
                                    new_element=CVSState()
                                    new_element.current_node_id=outgoing_node_id
                                    new_element.Label_waiting_time=pElement.Label_waiting_time+1
                                    sta_ii=get_sta_ii_of_waiting_node_train_id_ranked(n,k)  
                                    if new_element.Label_waiting_time<=g_train_list[k].max_dwell_tm_list[sta_ii]:
                                        new_element.Caculate_cost(1,outgoing_link_id,t,pElement,k)
                                        new_element.Update_node_time_link_sequence(pElement,n,t,outgoing_link_id)
                                        C_train_time_dependent_state_vector_lower_bound[k][next_t].update_state(new_element,1)
                                else:
                                    new_element=CVSState()
                                    new_element.current_node_id=outgoing_node_id
                                    new_element.Label_waiting_time=pElement.Label_waiting_time
                                    new_element.Caculate_cost(1,outgoing_link_id,t,pElement,k)
                                    new_element.Update_node_time_link_sequence(pElement,n,t,outgoing_link_id)
                                    C_train_time_dependent_state_vector_lower_bound[k][next_t].update_state(new_element,1)
                                    if new_element.current_node_id==current_train_dest_node_id:
                                        new_element.Add_end_n_and_end_t(current_train_dest_node_id,next_t)
                                        C_train_end_state_vector_lower_bound[k].CVSStateVector[0].append(new_element)
        if len(C_train_end_state_vector_lower_bound[k].CVSStateVector[0])==0:
            print('can not find lower_bound for train:{}'.format(k))
        else:
            C_train_end_state_vector_lower_bound[k].CVSStateVector[0].sort(key=lambda x:x.LabelCost_for_lr)
            g_train_list[k].node_sequence_lower_bound=C_train_end_state_vector_lower_bound[k].CVSStateVector[0][0].m_visit_node_sequence
            g_train_list[k].time_sequence_lower_bound=C_train_end_state_vector_lower_bound[k].CVSStateVector[0][0].m_visit_time_sequence
            g_train_list[k].link_sequence_lower_bound=C_train_end_state_vector_lower_bound[k].CVSStateVector[0][0].m_visit_link_sequence
            lower_bound[iteration_step]+=C_train_end_state_vector_lower_bound[k].CVSStateVector[0][0].LabelCost_for_lr                            
            C_train_end_state_vector_lower_bound[k].Reset()
    for l in range(1,g_number_of_links):
        for t in range(1,g_number_of_time_intervals):
            if g_link_list[l].link_type==2 or g_link_list[l].link_type==3 or g_link_list[l].link_type==4:
                lower_bound[iteration_step]-=g_link_list[l].time_dependent_LR_multiplier[t]
    for i in all_depart_nodes_of_all_stations:
        for t in range(1,g_number_of_time_intervals):
            lower_bound[iteration_step]-=g_node_list[i].time_dependent_LR_multiplier[t]
    for i in all_receive_nodes_of_all_stations:
        for t in range(1,g_number_of_time_intervals):
            lower_bound[iteration_step]-=g_node_list[i].time_dependent_LR_multiplier[t]
    print('22lower_bound[iteration_step]',lower_bound[iteration_step])


                
                                                                                       

def g_generate_and_calculate_upper_bound():
    global time_dependent_link_upper_delete_list
    global C_train_time_dependent_state_vector_upper_bound
    global C_train_end_state_vector_upper_bound
    already_sequence_indicator=0
    C_train_time_dependent_state_vector_upper_bound = [[None for t in range(1,g_number_of_time_intervals+1)] for k in range(1,g_number_of_trains+1)]
    C_train_end_state_vector_upper_bound = [None for k in range(1,g_number_of_trains+1)]
    #g_train_list_upper_bound=g_train_list
    g_train_list.sort(key=lambda x:len(x.conflict_trains_list))
    #g_train_list.sort(key=lambda x:x.current_iteration_upper_bound_priority)
    for k in range(1,g_number_of_trains):
        train_upper_bound_sequence[iteration_step].append(g_train_list[k].train_id)
    if iteration_step>0:
        for step in range(0,iteration_step):
            if operator.eq(train_upper_bound_sequence[iteration_step],train_upper_bound_sequence[step]):
                train_upper_bound_sequence[iteration_step]=[]
                exchange_train_1_position=random.randint(1,g_number_of_trains-1)
                if exchange_train_1_position==1:
                    g_train_list[exchange_train_1_position],g_train_list[2]=g_train_list[2],g_train_list[exchange_train_1_position]
                else:
                    g_train_list[exchange_train_1_position-1],g_train_list[exchange_train_1_position]=g_train_list[exchange_train_1_position],g_train_list[exchange_train_1_position-1]
                for k in range(1,g_number_of_trains):
                    train_upper_bound_sequence[iteration_step].append(g_train_list[k].train_id)
                for step in range(0,iteration_step):
                    if operator.eq(train_upper_bound_sequence[iteration_step],train_upper_bound_sequence[step]):
                        already_sequence_indicator=1
                break

    time_dependent_link_upper_delete_list=np.zeros([g_number_of_links,g_number_of_time_intervals]).astype(int)
    
    for k in range(1, g_number_of_trains):
        if iteration_step>0 and already_sequence_indicator==1: 
            upper_bound[iteration_step]=MAX_LABEL_COST
            break
        train_id_k=g_train_list[k].train_id
        if (len(g_train_list[k].conflict_trains_list)==0) and \
            (len(g_train_list[k].node_sequence)>0):
            for link_i in range(0,len(g_train_list[k].link_sequence)):
                l=g_train_list[k].link_sequence[link_i]
                if g_link_list[l].link_type>virtual_link_basic_number:
                    continue
                t=g_train_list[k].time_sequence[link_i]
                if  consider_waiting_arc_flag==1:
                    upper_bound[iteration_step]+=g_train_list[k].time_dependent_link_cost[l][t]
                else:
                    if g_link_list[l].link_type == 1 or(g_link_list[l].link_type==2) \
                                or (g_link_list[l].link_type == 4) or (g_link_list[l].link_type==3):
                        upper_bound[iteration_step]=upper_bound[iteration_step]+g_train_list[k].time_dependent_link_cost[l][t]
                if g_link_list[l].link_capacity==1:
                    time_dependent_link_upper_delete_list[l][t]=1
                if (g_link_list[l].link_type==1) or (g_link_list[l].link_type==2) or (g_link_list[l].link_type == 4):#rec and depart route
                    for i in range(0,len(g_link_list[l].IN_before_and_after_arcs_of_td_link[t])):
                        ll=g_link_list[l].IN_before_and_after_arcs_of_td_link[t][i][0]
                        tt=g_link_list[l].IN_before_and_after_arcs_of_td_link[t][i][1]
                        time_dependent_link_upper_delete_list[ll][tt]=1                            
                if (g_link_list[l].link_type == 3) : #wait arc
                    for tt in range(max(0,t-headway_track),
                                    min(t+1+headway_track,g_number_of_time_intervals)):
                        time_dependent_link_upper_delete_list[l][tt]=1
            g_train_list[k].node_sequence_upper_bound = g_train_list[k].node_sequence
            g_train_list[k].time_sequence_upper_bound = g_train_list[k].time_sequence
            g_train_list[k].link_sequence_upper_bound = g_train_list[k].link_sequence
    
        else:
        
            #get agent information
            current_train_orig_node_id = g_train_list[k].actual_orig_node_id
            current_train_dest_node_id = g_train_list[k].actual_dest_node_id
            original_tm_beginning = int(g_train_list[k].original_tm_beginning)

            
            #reset
            g_train_list[k].node_sequence_upper_bound = []
            g_train_list[k].time_sequence_upper_bound = []
            g_train_list[k].link_sequence_upper_bound = []

            # dynamic programming
            for t in range(0,g_number_of_time_intervals):
                C_train_time_dependent_state_vector_upper_bound[k][t] = C_time_indexed_state_vector()
                C_train_time_dependent_state_vector_upper_bound[k][t].Reset()
                C_train_time_dependent_state_vector_upper_bound[k][t].current_time=t
    

            C_train_end_state_vector_upper_bound[k]= C_time_indexed_state_vector()
            C_train_end_state_vector_upper_bound[k].Reset()
            #original_node
            element=CVSState()
            element.current_node_id=current_train_orig_node_id
            element.PrimalLabelCost=0
            C_train_time_dependent_state_vector_upper_bound[k][original_tm_beginning].update_state(element,0)


            for t in range(original_tm_beginning, g_number_of_time_intervals):
                C_train_time_dependent_state_vector_upper_bound[k][t].Sort(0)
                for w_index in range(min(BestKSize, len(C_train_time_dependent_state_vector_upper_bound[k][t].CVSStateVector[0]))):
                    pElement=C_train_time_dependent_state_vector_upper_bound[k][t].CVSStateVector[0][w_index]
                    n=pElement.current_node_id
                    if n==current_train_dest_node_id:
                        continue
                    for outgoing_link_id in g_node_list[n].outgoing_link_list:            
                        if (outgoing_link_id in g_train_list[k].available_links_list):
                            outgoing_node_id=g_link_list[outgoing_link_id].to_node_id
                            train_run_tm = g_link_list[outgoing_link_id].train_run_tm
                            if (time_dependent_link_blockage[outgoing_link_id][t] == 1) and (time_dependent_link_upper_delete_list[outgoing_link_id][t] == 0)  and g_train_list[k].available_arcs_list[outgoing_link_id][t]==1:
                                next_t=t+train_run_tm
                                if (next_t<g_number_of_time_intervals):
                                    if (n in g_all_stations_stop_nodes_list) and (not (outgoing_node_id in g_all_stations_stop_nodes_list)):
                                        sta_ii=get_sta_ii_of_waiting_node_train_id_ranked(n,k)                                        
                                        stop_pattern = g_train_list[k].stop_pattern_list[sta_ii]
                                        max_track_dwell_time=g_train_list[k].max_dwell_tm_list[sta_ii]
                                        if stop_pattern==1:                                   
                                            t_leave = g_train_list[k].plan_station_d_tm_list[sta_ii]
                                            min_train_dwell_time = g_train_list[k].min_dwell_tm_list[sta_ii]
                                            min_track_dwell_time = g_node_list[n].min_track_dwell_tm
                                            if (t>=t_leave) and pElement.Label_waiting_time>=min_train_dwell_time and \
                                                pElement.Label_waiting_time>=min_track_dwell_time and pElement.Label_waiting_time<=max_track_dwell_time:
                                                new_element=CVSState()
                                                new_element.current_node_id=outgoing_node_id
                                                new_element.Label_waiting_time=0
                                                new_element.Caculate_cost(0,outgoing_link_id,t,pElement,k)
                                                new_element.Update_node_time_link_sequence(pElement,n,t,outgoing_link_id)
                                                C_train_time_dependent_state_vector_upper_bound[k][next_t].update_state(new_element,0) 
                                                if new_element.current_node_id==current_train_dest_node_id:
                                                    new_element.Add_end_n_and_end_t(current_train_dest_node_id,next_t)
                                                    C_train_end_state_vector_upper_bound[k].CVSStateVector[0].append(new_element) 
                                        else:
                                            min_track_dwell_time = g_node_list[n].min_track_dwell_tm
                                            if pElement.Label_waiting_time>=min_track_dwell_time and pElement.Label_waiting_time<=max_track_dwell_time:
                                                new_element=CVSState()
                                                new_element.current_node_id=outgoing_node_id
                                                new_element.Label_waiting_time=0
                                                new_element.Caculate_cost(0,outgoing_link_id,t,pElement,k)
                                                new_element.Update_node_time_link_sequence(pElement,n,t,outgoing_link_id)
                                                C_train_time_dependent_state_vector_upper_bound[k][next_t].update_state(new_element,0)
                                                if new_element.current_node_id==current_train_dest_node_id:
                                                    new_element.Add_end_n_and_end_t(current_train_dest_node_id,next_t)
                                                    C_train_end_state_vector_upper_bound[k].CVSStateVector[0].append(new_element) 
                                    elif (n==outgoing_node_id) and n!=current_train_orig_node_id:
                                        new_element=CVSState()
                                        new_element.current_node_id=outgoing_node_id
                                        new_element.Label_waiting_time=pElement.Label_waiting_time+1
                                        sta_ii=get_sta_ii_of_waiting_node_train_id_ranked(n,k)  
                                        if new_element.Label_waiting_time<=g_train_list[k].max_dwell_tm_list[sta_ii]:
                                            new_element.Caculate_cost(0,outgoing_link_id,t,pElement,k)
                                            new_element.Update_node_time_link_sequence(pElement,n,t,outgoing_link_id)
                                            C_train_time_dependent_state_vector_upper_bound[k][next_t].update_state(new_element,0)
                                    else:
                                        new_element=CVSState()
                                        new_element.current_node_id=outgoing_node_id
                                        new_element.Label_waiting_time=pElement.Label_waiting_time
                                        new_element.Caculate_cost(0,outgoing_link_id,t,pElement,k)
                                        new_element.Update_node_time_link_sequence(pElement,n,t,outgoing_link_id)
                                        C_train_time_dependent_state_vector_upper_bound[k][next_t].update_state(new_element,0)
                                        if new_element.current_node_id==current_train_dest_node_id:
                                            new_element.Add_end_n_and_end_t(current_train_dest_node_id,next_t)
                                            C_train_end_state_vector_upper_bound[k].CVSStateVector[0].append(new_element)

            if len(C_train_end_state_vector_upper_bound[k].CVSStateVector[0])==0:
                print('can not find upper_bound for train:{}'.format(train_id_k))
                upper_bound[iteration_step]=upper_bound[iteration_step]+virtual_vehicle_fee
            else:
                C_train_end_state_vector_upper_bound[k].CVSStateVector[0].sort(key=lambda x:x.PrimalLabelCost)
                g_train_list[k].node_sequence_upper_bound=C_train_end_state_vector_upper_bound[k].CVSStateVector[0][0].m_visit_node_sequence
                g_train_list[k].time_sequence_upper_bound=C_train_end_state_vector_upper_bound[k].CVSStateVector[0][0].m_visit_time_sequence
                g_train_list[k].link_sequence_upper_bound=C_train_end_state_vector_upper_bound[k].CVSStateVector[0][0].m_visit_link_sequence
                C_train_end_state_vector_upper_bound[k].Reset()
                for link_i in range(0,len(g_train_list[k].link_sequence_upper_bound)):
                    l=g_train_list[k].link_sequence_upper_bound[link_i]
                    if g_link_list[l].link_type>virtual_link_basic_number:
                        continue
                    t=g_train_list[k].time_sequence_upper_bound[link_i]
                    if consider_waiting_arc_flag==1:
                        upper_bound[iteration_step]+=g_train_list[k].time_dependent_link_cost[l][t]
                    else:
                        if g_link_list[l].link_type == 1 or(g_link_list[l].link_type==2) \
                            or (g_link_list[l].link_type == 4) or (g_link_list[l].link_type==3):
                            upper_bound[iteration_step]=upper_bound[iteration_step]+g_train_list[k].time_dependent_link_cost[l][t]
                    if g_link_list[l].link_capacity==1:
                        time_dependent_link_upper_delete_list[l][t]=1
                    if (g_link_list[l].link_type==1) or (g_link_list[l].link_type==2) or\
                        (g_link_list[l].link_type == 4) :#rec and depart route
                        for i in range(0,len(g_link_list[l].IN_before_and_after_arcs_of_td_link[t])):
                            ll=g_link_list[l].IN_before_and_after_arcs_of_td_link[t][i][0]
                            tt=g_link_list[l].IN_before_and_after_arcs_of_td_link[t][i][1]
                            time_dependent_link_upper_delete_list[ll][tt]=1
                    if (g_link_list[l].link_type == 3):#wait arc
                        for tt in range(max(1,t-headway_track),
                                        min(t+1+headway_track,g_number_of_time_intervals)):
                            time_dependent_link_upper_delete_list[l][tt]=1 

                
    g_train_list.sort(key=lambda x:x.train_id)
    print('22upper_bound',upper_bound[iteration_step])
    return()
    


def g_update_multiplier_2():
    global roh
    global u
    violation1=0
    violation2=0
    for l in range(1, g_number_of_links):
        if (g_link_list[l].link_type>=virtual_link_basic_number):
            break
        for t in range(1, g_number_of_time_intervals):
            if (g_link_list[l].link_type == 1) or (g_link_list[l].link_type == 2) or (g_link_list[l].link_type == 4):#rec and depart route
                
                sum_of_others_1_2_4=0
                for i in range(0,len(g_link_list[l].IN_after_arcs_of_td_link[t])):
                    ll=g_link_list[l].IN_after_arcs_of_td_link[t][i][0]
                    tt=g_link_list[l].IN_after_arcs_of_td_link[t][i][1]
                    sum_of_others_1_2_4+=time_dependent_link_volume_for_trains[ll][tt]
                g_link_list[l].time_dependent_LR_multiplier[t]+=(sum_of_others_1_2_4-1)*roh#**
                #violation1=sum_of_others_1_2_4-1
                #violation1=max(0,violation1)
            if (g_link_list[l].link_type == 3):
                g_link_list[l].time_dependent_LR_multiplier[t]+=(time_dependent_link_volume_for_trains[l][t]*u)#**
                g_link_list[l].time_dependent_LR_multiplier[t]+=(time_dependent_implied_waiting_link_flag_for_trains[l][t]*u)#**
                g_link_list[l].time_dependent_LR_multiplier[t]-=u#**
                #violation2=time_dependent_link_volume_for_trains[l][t]+time_dependent_implied_waiting_link_flag_for_trains[l][t]-1
                #violation2=max(0,violation2)
            g_link_list[l].time_dependent_LR_multiplier[t]=max(0,g_link_list[l].time_dependent_LR_multiplier[t])
    if violation1*violation1>=v_roh*violation1:
        roh=roh+d_roh
    if violation2*violation2>=v_u*violation2:
        u=u+d_u
        
def g_update_multiplier_1():
    for l in range(1, g_number_of_links):
        if (g_link_list[l].link_type>=virtual_link_basic_number):
            break
        if (g_link_list[l].link_type==1):#depart
            continue
        for t in range(1, g_number_of_time_intervals):
            if (g_link_list[l].link_type == 4):#section route
                sum_of_others_4=0
                for i in range(0,len(g_link_list[l].IN_after_arcs_of_td_link[t])):
                    ll=g_link_list[l].IN_after_arcs_of_td_link[t][i][0]
                    tt=g_link_list[l].IN_after_arcs_of_td_link[t][i][1]
                    sum_of_others_4+=time_dependent_link_volume_for_trains[ll][tt]
                g_link_list[l].time_dependent_LR_multiplier[t]+=(sum_of_others_4-1)*roh4#**
            if (g_link_list[l].link_type == 3):#waiting arc
                g_link_list[l].time_dependent_LR_multiplier[t]+=(time_dependent_link_volume_for_trains[l][t]+time_dependent_implied_waiting_link_flag_for_trains[l][t]-1)*roh5
            if (g_link_list[l].link_type == 2):#receiving
                i_l_from_node=g_link_list[l].from_node_id
                j_l_to_node=g_link_list[l].to_node_id
                sum_of_others_2=0
                for i in range(0,len(g_link_list[l].IN_after_arcs_of_td_link[t])):
                    ll=g_link_list[l].IN_after_arcs_of_td_link[t][i][0]
                    ii_ll_from_node=g_link_list[ll].from_node_id
                    jj_ll_to_node=g_link_list[ll].to_node_id
                    if not((ii_ll_from_node==i_l_from_node) and (jj_ll_to_node!=j_l_to_node)):
                        tt=g_link_list[l].IN_after_arcs_of_td_link[t][i][1]
                        sum_of_others_2+=time_dependent_link_volume_for_trains[ll][tt]
                g_link_list[l].time_dependent_LR_multiplier[t]+=(sum_of_others_2-1)*roh3#**
            g_link_list[l].time_dependent_LR_multiplier[t]=max(0,g_link_list[l].time_dependent_LR_multiplier[t])
    for i in all_receive_nodes_of_all_stations:
        receive_route_run_tm_of_i=g_node_list[i].station_route_run_tm
        for t in range(1, g_number_of_time_intervals):
            sum_of_others_receive_node=0
            for tt in range(t,min(g_number_of_time_intervals,t+receive_route_run_tm_of_i+headway_route)):
                for l in g_node_list[i].outgoing_link_list:
                    if g_link_list[l].link_type<virtual_link_basic_number:
                        sum_of_others_receive_node+=time_dependent_link_volume_for_trains[l][tt]
            g_node_list[i].time_dependent_LR_multiplier[t]+=(sum_of_others_receive_node-1)*roh1
            g_node_list[i].time_dependent_LR_multiplier[t]=max(0,g_node_list[i].time_dependent_LR_multiplier[t])
    for j in all_depart_nodes_of_all_stations:
        depart_route_run_tm_of_j=g_node_list[j].station_route_run_tm
        for tuo in range(1,g_number_of_time_intervals):
            sum_of_others_depart_node=0
            for tuotuo in range(tuo, min(g_number_of_time_intervals,tuo+depart_route_run_tm_of_j+headway_route)):
                t=tuotuo-depart_route_run_tm_of_j
                if t>=1:
                    for l in g_node_list[j].ingoing_link_list:
                        if g_link_list[l].link_type<virtual_link_basic_number:
                            sum_of_others_depart_node+=time_dependent_link_volume_for_trains[l][t]
            g_node_list[j].time_dependent_LR_multiplier[tuo]+=(sum_of_others_depart_node-1)*roh2
            g_node_list[j].time_dependent_LR_multiplier[tuo]=max(0,g_node_list[j].time_dependent_LR_multiplier[tuo])

def update_best_result():
    if (iteration_step == 0):
        for k in range(1, g_number_of_trains):       
            g_train_list[k].best_node_sequence =copy.copy(g_train_list[k].node_sequence)
            g_train_list[k].best_time_sequence = copy.copy(g_train_list[k].time_sequence)
            g_train_list[k].best_link_sequence = copy.copy(g_train_list[k].link_sequence)      
            g_train_list[k].best_node_sequence_upper_bound = copy.copy(g_train_list[k].node_sequence_upper_bound)
            g_train_list[k].best_time_sequence_upper_bound = copy.copy(g_train_list[k].time_sequence_upper_bound)
            g_train_list[k].best_link_sequence_upper_bound = copy.copy(g_train_list[k].link_sequence_upper_bound)
            g_train_list[k].best_node_sequence_lower_bound = copy.copy(g_train_list[k].node_sequence_lower_bound)
            g_train_list[k].best_time_sequence_lower_bound = copy.copy(g_train_list[k].time_sequence_lower_bound)
            g_train_list[k].best_link_sequence_lower_bound = copy.copy(g_train_list[k].link_sequence_lower_bound) 
        print("iteration_step{},upper_bound{}".format(iteration_step,upper_bound[iteration_step]))
        print("iteration_step{},lower_bound{}".format(iteration_step,lower_bound[iteration_step]))
        print("iteration_step{},train_result{}".format(iteration_step,train_result[iteration_step]))

    if (iteration_step >= 1):
        if (upper_bound[iteration_step] <= upper_bound[iteration_step - 1]):
            for k in range(1, g_number_of_trains):
                g_train_list[k].best_node_sequence_upper_bound = copy.copy(g_train_list[k].node_sequence_upper_bound)
                g_train_list[k].best_time_sequence_upper_bound = copy.copy(g_train_list[k].time_sequence_upper_bound)
                g_train_list[k].best_link_sequence_upper_bound = copy.copy(g_train_list[k].link_sequence_upper_bound)
            upper_bound[iteration_step]=upper_bound[iteration_step]
        else:
            for k in range(1, g_number_of_trains):
                g_train_list[k].best_node_sequence_upper_bound = copy.copy(g_train_list[k].best_node_sequence_upper_bound)
                g_train_list[k].best_time_sequence_upper_bound = copy.copy(g_train_list[k].best_time_sequence_upper_bound)
                g_train_list[k].best_link_sequence_upper_bound = copy.copy(g_train_list[k].best_link_sequence_upper_bound)                
            upper_bound[iteration_step]=upper_bound[iteration_step-1]
        print("iteration_step{},upper_bound{}".format(iteration_step,upper_bound[iteration_step]))
        
        if lower_bound[iteration_step]>lower_bound[iteration_step-1]:
            for k in range(1, g_number_of_trains):
                g_train_list[k].best_node_sequence_lower_bound = copy.copy(g_train_list[k].node_sequence_lower_bound)
                g_train_list[k].best_time_sequence_lower_bound = copy.copy(g_train_list[k].time_sequence_lower_bound)
                g_train_list[k].best_link_sequence_lower_bound = copy.copy(g_train_list[k].link_sequence_lower_bound)            
            lower_bound[iteration_step]=lower_bound[iteration_step]
        else:
            for k in range(1, g_number_of_trains):
                g_train_list[k].best_node_sequence_lower_bound = copy.copy(g_train_list[k].best_node_sequence_lower_bound)
                g_train_list[k].best_time_sequence_lower_bound = copy.copy(g_train_list[k].best_time_sequence_lower_bound)
                g_train_list[k].best_link_sequence_lower_bound = copy.copy(g_train_list[k].best_link_sequence_lower_bound)            
            lower_bound[iteration_step]=lower_bound[iteration_step-1]
        print("iteration_step{},lower_bound{}".format(iteration_step,lower_bound[iteration_step]))
        
        for k in range(1, g_number_of_trains):
            g_train_list[k].best_node_sequence = copy.copy(g_train_list[k].node_sequence)
            g_train_list[k].best_time_sequence = copy.copy(g_train_list[k].time_sequence)
            g_train_list[k].best_link_sequence = copy.copy(g_train_list[k].link_sequence)
        print("iteration_step{},train_result{}".format(iteration_step,train_result[iteration_step]))

    
def g_write_output_data(): 
    # output train results
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'output_train_results'
    sheet['A1'] = 'train_id'
    sheet['B1'] = 'best_node_sequence'
    sheet['C1'] = 'best_time_sequence'
    sheet['D1'] = 'best_link_sequence'
    for k in range(1, g_number_of_trains):
        row = k + 1
        best_node_sequence = ";".join(str(node) for node in g_train_list[k].best_node_sequence)
        best_time_sequence = ";".join(str(time) for time in g_train_list[k].best_time_sequence)
        best_link_sequence = ";".join(str(link) for link in g_train_list[k].best_link_sequence) 
        sheet.cell(row=row, column=1, value=k)
        sheet.cell(row=row, column=2, value=best_node_sequence)
        sheet.cell(row=row, column=3, value=best_time_sequence)
        sheet.cell(row=row, column=4, value=best_link_sequence)

    workbook.save('output_train_results.xlsx')      
    

    # output train upper bound
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'output_train_upper_results'
    sheet['A1'] = 'train_id'
    sheet['B1'] = 'best_node_sequence_upper_bound'
    sheet['C1'] = 'best_time_sequence_upper_bound'
    sheet['D1'] = 'best_link_sequence_upper_bound'
    for k in range(1, g_number_of_trains):
        row = k + 1
        best_node_sequence_upper_bound = ";".join(str(node) for node in g_train_list[k].best_node_sequence_upper_bound)
        best_time_sequence_upper_bound = ";".join(str(time) for time in g_train_list[k].best_time_sequence_upper_bound)
        best_link_sequence_upper_bound = ";".join(str(link) for link in g_train_list[k].best_link_sequence_upper_bound)            
        sheet.cell(row=row, column=1, value=k)
        sheet.cell(row=row, column=2, value=best_node_sequence_upper_bound)
        sheet.cell(row=row, column=3, value=best_time_sequence_upper_bound)
        sheet.cell(row=row, column=4, value=best_link_sequence_upper_bound)
    workbook.save('output_train_upper_results.xlsx')
    
    # output train lower bound
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'output_train_lower_results'
    sheet['A1'] = 'train_id'
    sheet['B1'] = 'best_node_sequence_lower_bound'
    sheet['C1'] = 'best_time_sequence_lower_bound'
    sheet['D1'] = 'best_link_sequence_lower_bound'
    for k in range(1, g_number_of_trains):
        row = k + 1
        best_node_sequence_lower_bound = ";".join(str(node) for node in g_train_list[k].best_node_sequence_lower_bound)
        best_time_sequence_lower_bound = ";".join(str(time) for time in g_train_list[k].best_time_sequence_lower_bound)
        best_link_sequence_lower_bound = ";".join(str(link) for link in g_train_list[k].best_link_sequence_lower_bound)
        sheet.cell(row=row, column=1, value=k)
        sheet.cell(row=row, column=2, value=best_node_sequence_lower_bound)
        sheet.cell(row=row, column=3, value=best_time_sequence_lower_bound)
        sheet.cell(row=row, column=4, value=best_link_sequence_lower_bound)
    workbook.save('output_train_lower_results.xlsx')
    
def IN_before_arcs_of_td_link(l,t):
    if (g_link_list[l].link_type == 1) or (g_link_list[l].link_type == 2):
        for ll in g_link_list[l].conflict_link_list:
            route_run_tm=g_link_list[l].train_run_tm
            for tt in range(max(1,t-route_run_tm-headway_route+1),
                            min(t+1,g_number_of_time_intervals)):
                g_link_list[l].IN_before_arcs_of_td_link[t].append([ll,tt])
    if (g_link_list[l].link_type == 4):
        for ll in g_link_list[l].conflict_link_list:
            headway_section = max(headway_station_a,headway_station_b)
            for tt in range(max(1,t-headway_section+1),
                            min(t+1,g_number_of_time_intervals)):
                g_link_list[l].IN_before_arcs_of_td_link[t].append([ll,tt])

def IN_after_arcs_of_td_link(l,t):
    if (g_link_list[l].link_type == 1) or (g_link_list[l].link_type == 2):
        for ll in g_link_list[l].conflict_link_list:
            route_run_tm=g_link_list[l].train_run_tm
            for tt in range(t, min(t + route_run_tm+headway_route,g_number_of_time_intervals)):
                g_link_list[l].IN_after_arcs_of_td_link[t].append([ll,tt])
    if (g_link_list[l].link_type == 4):
        for ll in g_link_list[l].conflict_link_list:
            headway_section = max(headway_station_a,headway_station_b)
            for tt in range(t, min(t + headway_section,g_number_of_time_intervals)):
                g_link_list[l].IN_after_arcs_of_td_link[t].append([ll,tt])

                
def IN_before_and_after_arcs_of_td_link(l,t):
    if (g_link_list[l].link_type == 1) or (g_link_list[l].link_type == 2):
        for ll in g_link_list[l].conflict_link_list:
            route_run_tm=g_link_list[l].train_run_tm
            for tt in range(max(1,t-route_run_tm-headway_route+1), min(t + route_run_tm+headway_route,g_number_of_time_intervals)):
                g_link_list[l].IN_before_and_after_arcs_of_td_link[t].append([ll,tt])
    if (g_link_list[l].link_type == 4):
        for ll in g_link_list[l].conflict_link_list:
            headway_section = max(headway_station_a,headway_station_b)
            for tt in range(max(1,t-headway_section+1), min(t + headway_section,g_number_of_time_intervals)):
                g_link_list[l].IN_before_and_after_arcs_of_td_link[t].append([ll,tt])  

def IN_after_arcs_of_td_receive_node(i,t):
    for l in g_node_list[i].outgoing_link_list:
        if g_link_list[l].link_type>=virtual_link_basic_number:
            continue
        route_run_tm=g_link_list[l].train_run_tm
        for tt in range(t,min(t+route_run_tm+headway_route,g_number_of_time_intervals)):
            g_node_list[i].IN_after_arcs_of_td_receive_node[t].append([l,tt])
            
def IN_before_and_after_arcs_of_td_receive_node(i,t):
    for l in g_node_list[i].outgoing_link_list:
        if g_link_list[l].link_type>=virtual_link_basic_number:
            continue
        route_run_tm=g_link_list[l].train_run_tm
        for tt in range(max(1,t-route_run_tm-headway_route+1),min(t+route_run_tm+headway_route,g_number_of_time_intervals)):
            g_node_list[i].IN_before_and_after_arcs_of_td_receive_node[t].append([l,tt])
    
def IN_after_arcs_of_td_depart_node(j,tuo):
    for l in g_node_list[j].ingoing_link_list:
        if g_link_list[j].link_type>=virtual_link_basic_number:
            continue
        route_run_tm=g_link_list[l].train_run_tm
        t=tuo-route_run_tm
        if t>=1:
            for tt in range(max(t,1),min(tuo+headway_route,g_number_of_time_intervals)):
                g_node_list[j].IN_after_arcs_of_td_depart_node[tuo].append([l,tt])
            
def IN_before_and_after_arcs_of_td_depart_node(j,tuo):
    for l in g_node_list[j].ingoing_link_list:
        if g_link_list[l].link_type>=virtual_link_basic_number:
            continue
        route_run_tm=g_link_list[l].train_run_tm
        t=tuo-route_run_tm
        if t>=1:
            for tt in range(max(1,t-route_run_tm-headway_route+1),min(tuo+headway_route,g_number_of_time_intervals)):
                g_node_list[j].IN_before_and_after_arcs_of_td_depart_node[tuo].append([l,tt])

            
if __name__=='__main__':
    print('Reading data......')
    g_number_of_time_intervals =80#need change
    all_virtual_wt_node=[]
    virtual_ori_wt_arc_cost=4#need change
    waiting_link_cost=3#need change
    max_t_implicit_occupation=1
    virtual_dest_wt_arc_cost=0
    T_orig_eariler=0
    T_orig_later=20
    T_dest_later=20
    all_depart_nodes_of_all_stations=[]
    all_receive_nodes_of_all_stations=[]
    initial_multiplier=0#need change
    virtual_link_basic_number=400
    g_ReadInputData()     
    MAX_LABEL_COST = 1000000
    BestKSize=1000
    maximum_external_iteration_step =15
    headway_station_a = 3#need change
    headway_station_b = 3#need change
    headway_section = max(headway_station_a,headway_station_b)
    headway_route = int(1)#need change
    headway_track = int(1)#need change
    #test
    v_roh=0.5
    d_roh=0.5
    v_u=0.5
    d_u=0.5
    #test_over
    consider_waiting_arc_flag=1

    #initial lower and upper bound 
    lower_bound = np.zeros([maximum_external_iteration_step + 1])
    upper_bound = np.zeros([maximum_external_iteration_step + 1])
    train_result = np.zeros([maximum_external_iteration_step + 1])
    result_feasibility = np.zeros([maximum_external_iteration_step + 1])
    train_upper_bound_sequence=[[]for step in range(0,maximum_external_iteration_step+1)]
    #reset the (waiting) link volume and best_result of each train
    g_state_list=[[[]for t in range(1,g_number_of_time_intervals+1)] for k in range(1,g_number_of_trains+1)]
    for k in range(1,g_number_of_trains):
        g_train_list[k].time_dependent_link_volume = [[0 for t in range(1, g_number_of_time_intervals + 1)] for l in
                                                      range(1, g_number_of_links + 1)]
        g_train_list[k].time_dependent_implied_waiting_link_flag = [[0 for t in range(1, g_number_of_time_intervals + 1)] for l in
                                                      range(1, g_number_of_links + 1)]
        g_train_list[k].best_time_dependent_link_volume = [[0 for t in range(1, g_number_of_time_intervals + 1)] for l in 
                                                      range(1, g_number_of_links + 1)]
        g_train_list[k].train_free_flow_time=g_train_list[k].plan_dest_arrival_tm-g_train_list[k].original_tm_beginning

    #initial quadratic penalty and lr multipier and arc cost
    roh1_0 =2#need change
    roh2_0 =2#need change
    roh3_0 =2#need change
    roh4_0 =2#need change
    roh5_0 =2#need change
    print('roh1,roh2,roh3,roh4,roh5',roh1_0,roh2_0,roh3_0,roh4_0,roh5_0)
    g_initialize_multiplier() 
    g_initialize_cost()
    
    #virtual parameter
    virtual_vehicle_fee =g_number_of_time_intervals*10

            
    #define available td blockage=1, otherwise =-1
    time_dependent_link_blockage()#float
    for k in range(1,g_number_of_trains):
        train_time_dependent_link_cost(k)    
    # initial link volume for all trains
    time_dependent_link_volume_for_trains = [[0 for t in range(1, g_number_of_time_intervals + 1)] for l in range(1, g_number_of_links + 1)]
    time_dependent_implied_waiting_link_flag_for_trains = [[0 for t in range(1, g_number_of_time_intervals + 1)] for l in range(1, g_number_of_links + 1)]
    time_dependent_link_all_occupation= [[0 for t in range(1, g_number_of_time_intervals + 1)] for l in range(1, g_number_of_links + 1)]
    roh1 = roh1_0
    roh2 = roh2_0
    roh3 = roh3_0
    roh4 = roh4_0
    roh5 = roh5_0

    #pre-calculate each arc conflicting arcs of td links before
    #initial
    for l in range(1,g_number_of_links):
        for t in range(1,g_number_of_time_intervals):
            g_link_list[l].IN_before_arcs_of_td_link=[[]for t in range(1, g_number_of_time_intervals + 1)]           
            g_link_list[l].IN_after_arcs_of_td_link=[[]for t in range(1, g_number_of_time_intervals + 1)]            
            g_link_list[l].IN_before_and_after_arcs_of_td_link=[[]for t in range(1, g_number_of_time_intervals + 1)]
    #pre-calculate                
    for l in range(1,g_number_of_links):
        for t in range(1,g_number_of_time_intervals):            
            IN_before_arcs_of_td_link(l,t)
            IN_after_arcs_of_td_link(l,t)
            IN_before_and_after_arcs_of_td_link(l,t)
            
    #pre-calculate each arc conflicting arcs of td receive nodes before
    #initial
    for i in range(1,g_number_of_nodes):
        if g_node_list[i].node_type==1 or g_node_list[i].node_type==2:
            for t in range(1,g_number_of_time_intervals):
                g_node_list[i].IN_after_arcs_of_td_receive_node=[[]for t in range(1, g_number_of_time_intervals + 1)]            
                g_node_list[i].IN_before_and_after_arcs_of_td_receive_node=[[]for t in range(1, g_number_of_time_intervals + 1)]
    #pre-calculate                
    for i in range(1,g_number_of_nodes):
        if g_node_list[i].node_type==1 or g_node_list[i].node_type==2:
            for t in range(1,g_number_of_time_intervals):
                IN_after_arcs_of_td_receive_node(i,t)
                IN_before_and_after_arcs_of_td_receive_node(i,t)  
    
    #pre-calculate each arc conflicting arcs of td depart nodes before
    #initial
    for j in range(1,g_number_of_nodes):
        if g_node_list[j].node_type==3 or g_node_list[j].node_type==4:
            for tuo in range(1,g_number_of_time_intervals):
                g_node_list[j].IN_after_arcs_of_td_depart_node=[[]for tuo in range(1, g_number_of_time_intervals + 1)]            
                g_node_list[j].IN_before_and_after_arcs_of_td_depart_node=[[]for tuo in range(1, g_number_of_time_intervals + 1)]
    #pre-calculate                
    for j in range(1,g_number_of_nodes):
        if g_node_list[j].node_type==3 or g_node_list[j].node_type==4:
            for tuo in range(1,g_number_of_time_intervals):
                IN_after_arcs_of_td_depart_node(j,tuo)
                IN_before_and_after_arcs_of_td_depart_node(j,tuo)  
         
    
    time_start = time.time()
    for iteration_step in range(0, maximum_external_iteration_step):
        print('current_iteration_step:',iteration_step)
        calculate_LR_for_current_iteration_4()
        g_calcualte_lower_bound_2()
        g_time_dependent_dynamic_programming_for_trains() 
        if result_feasibility[iteration_step]==0:
            g_generate_and_calculate_upper_bound()
        else:
            cancel_indicator=0
            for k in range(1,g_number_of_trains):
                if len(g_train_list[k].link_sequence)==0:
                    cancel_indicator+=1
            if cancel_indicator!=0:
                g_generate_and_calculate_upper_bound()
            else:
                for k in range(1,g_number_of_trains):
                    train_upper_bound_sequence[iteration_step].append(g_train_list[k].train_id)
                    g_train_list[k].link_sequence_upper_bound=g_train_list[k].link_sequence
                    g_train_list[k].time_sequence_upper_bound=g_train_list[k].time_sequence
                    g_train_list[k].node_sequence_upper_bound=g_train_list[k].node_sequence
                    upper_bound[iteration_step]=train_result[iteration_step]
        if upper_bound[iteration_step]==lower_bound[iteration_step]:
            update_best_result()
            break       
        time_now=time.time()
        if time_now-time_start>3600:
            update_best_result()
            break
        g_update_multiplier_1()#calculate directly after
        update_best_result()
    time_end = time.time()
    g_write_output_data()
    print('time cost', time_end - time_start)




                      
        